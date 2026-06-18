import json
import math
import os
import sys
import warnings

import openpyxl
import pandas as pd
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from rich.console import Console
from rich.prompt import Confirm

from recipe_automation.core.config import settings

# NOT: do_match_depo burada import edilmiyor — circular import'u önlemek için
# fonksiyon içinde lazy import yapılıyor (bkz. run_capacity_analysis içindeki for döngüsü)

console = Console()


def norm_col(text: str) -> str:
    if not text:
        return ""
    replacements = {
        "ı": "i",
        "İ": "i",
        "I": "i",
        "ş": "s",
        "Ş": "s",
        "ğ": "g",
        "Ğ": "g",
        "ü": "u",
        "Ü": "u",
        "ö": "o",
        "Ö": "o",
        "ç": "c",
        "Ç": "c",
    }
    val = str(text)
    for tr, eng in replacements.items():
        val = val.replace(tr, eng)
    return val.strip().lower()


def safe_kod(val) -> str:
    s = str(val).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s


def safe_float(val) -> float:
    if pd.isna(val):
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", ".")
    try:
        return float(s)
    except:
        return 0.0


def run_capacity_analysis():
    console.print("\n[bold magenta]=== KAPASİTE VE FİZİBİLİTE ANALİZİ ===[/bold magenta]")

    db_dir = os.path.join(os.getcwd(), settings.db_dir_name)

    # 1. Grup Seçimi
    json_path = os.path.join(db_dir, "operasyon_gruplari.json")

    groups_data = {}
    if os.path.exists(json_path):
        with open(json_path, encoding="utf-8") as f:
            groups_data = json.load(f)

    if not groups_data:
        console.print("[red]operasyon_gruplari.json bulunamadı.[/red]")
        return

    group_keys = list(groups_data.keys())
    console.print("\n[bold cyan]Soru:[/bold cyan] Hangi grup için analiz yapmak istersiniz?")
    for i, g in enumerate(group_keys, 1):
        console.print(f"  [{i}] {g}")
    console.print(f"  [{len(group_keys) + 1}] TÜM PARÇALAR (Filtresiz Genel Analiz)")

    choice = console.input(f"Seçiminiz [1-{len(group_keys)+1}]: ").strip()
    try:
        idx = int(choice) - 1
        if 0 <= idx < len(group_keys):
            selected_group_name = group_keys[idx]
            target_operations = set([x.strip().upper() for x in groups_data[selected_group_name]])
        elif idx == len(group_keys):
            selected_group_name = "TÜM_PARÇALAR"
            target_operations = None
        else:
            console.print("[red]Geçersiz seçim![/red]")
            return
    except ValueError:
        console.print("[red]Lütfen bir rakam girin![/red]")
        return

    # 3. Üretilecek Set (Adet) Seçimi
    while True:
        try:
            adet_input = console.input(
                "\n[bold cyan]Soru:[/bold cyan] Bu reçeteden TOPLAM KAÇ ADET (set) üretmek istiyorsunuz? "
            ).strip()
            hedef_adet = int(adet_input)
            if hedef_adet <= 0:
                console.print("[yellow]Lütfen 0'dan büyük bir sayı girin.[/yellow]")
                continue
            break
        except ValueError:
            console.print("[red]Lütfen geçerli bir sayı girin![/red]")

    # 3.2 Stok Tipi Seçimi (Kullanılabilir vs Fiziki vs Her İkisi)
    console.print(
        "\n[bold cyan]Soru:[/bold cyan] Net üretim ihtiyacı (Üretilecek Miktar) hesaplanırken hangi stok tipini baz almak istersiniz?"
    )
    console.print("  [1] Kullanılabilir Stok (Fiziki Stok - Rezerveler) (Önerilen)")
    console.print("  [2] Fiziki Stok (Rezerveler yoksayılır)")
    console.print("  [3] Her İkisi de (Ayrı dosyalar halinde üretilir)")
    stok_secimi = console.input("Seçiminiz [1-3] (1): ").strip()

    if stok_secimi == "3":
        stock_bases = ["kullanilabilir", "fiziki"]
    elif stok_secimi == "2":
        stock_bases = ["fiziki"]
    else:
        stock_bases = ["kullanilabilir"]

    # 4. Hariç Tutma Sorusu
    haric_tut = Confirm.ask(
        "[bold yellow]Soru:[/bold yellow] Merdane, boru, kaynak vb. parçalar (haric_tutulacak_parcalar.json listesindeki) yoksayılsın mı?",
        default=False,
    )

    # 5. Dosyaları Bul
    input_dir = os.getcwd()

    # Sürükle bırak argümanı var mı?
    if len(sys.argv) > 1 and os.path.exists(sys.argv[1]):
        dragged_path = sys.argv[1].strip('"').strip("'")
        if os.path.isfile(dragged_path) and dragged_path.lower().endswith((".xlsx", ".xls")):
            file_path = dragged_path
            file_name = os.path.basename(dragged_path)
            input_dir = os.path.dirname(dragged_path)
        else:
            input_dir = (
                dragged_path if os.path.isdir(dragged_path) else os.path.dirname(dragged_path)
            )
            excel_files = [
                f
                for f in os.listdir(input_dir)
                if f.endswith((".xlsx", ".xls"))
                and not f.startswith(("~$", "ESLESTI_", "TOPLU_", "ANALIZ_"))
                and f
                not in ["TumRotaBilgileri.xlsx", "ReceteTumRotaListe.xlsx", "StokListesi.xlsx"]
            ]
            if not excel_files:
                console.print(
                    "[red]Sürüklediğiniz klasörde analiz edilecek Reçete Excel'i bulunamadı![/red]"
                )
                return
            file_name = excel_files[0]
            file_path = os.path.join(input_dir, file_name)
    else:
        excel_files = [
            f
            for f in os.listdir(input_dir)
            if f.endswith((".xlsx", ".xls"))
            and not f.startswith(("~$", "ESLESTI_", "TOPLU_", "ANALIZ_"))
            and f not in ["TumRotaBilgileri.xlsx", "ReceteTumRotaListe.xlsx", "StokListesi.xlsx"]
        ]

        if not excel_files:
            console.print("[red]Klasörde analiz edilecek kaynak dosya bulunamadı![/red]")
            return

        file_name = excel_files[0]
        file_path = os.path.join(input_dir, file_name)

    base_file_name = os.path.splitext(file_name)[0]
    console.print(f"\n[cyan]Analiz ediliyor: {file_name} ...[/cyan]")

    # 6. Kaynak Dosyayı Oku
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        try:
            df_source = pd.read_excel(file_path, header=None)
        except Exception as e:
            console.print(f"[red]Dosya okunurken hata: {e}[/red]")
            return

    header_idx = -1
    for i, row in df_source.iterrows():
        row_str = " ".join([str(x).lower() for x in row.values])
        if "kod" in row_str and ("malzeme" in row_str or "ad" in row_str):
            header_idx = i
            break
    if header_idx == -1:
        console.print("[red]Dosya başlıkları bulunamadı.[/red]")
        return

    df = df_source.iloc[header_idx + 1 :].copy()
    df.columns = df_source.iloc[header_idx]
    df = df.dropna(how="all")

    # Sütunları ve miktar başlıklarını bul
    kod_col = None
    sira_col = None

    for c in df.columns:
        nc = norm_col(str(c))
        if nc in ["kod", "malzeme kodu"]:
            kod_col = c
        if nc in ["sıra no", "sira no"]:
            sira_col = c

    miktar_col_kull = None
    for target in ["Rezerve Edilecek Miktar"]:
        for c in df.columns:
            if target.lower() in str(c).lower():
                miktar_col_kull = c
                break
        if miktar_col_kull:
            break

    miktar_col_fiziki = None
    for target in ["Çarpılmış Miktar", "Carpilmis Miktar", "Miktar"]:
        for c in df.columns:
            if target.lower() in str(c).lower():
                miktar_col_fiziki = c
                break
        if miktar_col_fiziki:
            break

    if not miktar_col_kull:
        miktar_col_kull = miktar_col_fiziki
    if not miktar_col_fiziki:
        miktar_col_fiziki = miktar_col_kull

    if not kod_col or not miktar_col_fiziki:
        console.print(
            "[red]Dosyada Kod veya Çarpılmış Miktar/Rezerve Edilecek Miktar sütunu bulunamadı![/red]"
        )
        return

    op_cols = []
    for c in df.columns:
        if (
            "operasyon" in norm_col(str(c))
            and "adi" not in norm_col(str(c))
            and "sira" not in norm_col(str(c))
        ):
            op_cols.append(c)

    # Orijinal Reçetedeki Ana Montaj Adetini Bul (Normalizasyon için)
    root_qty = 1.0
    if len(df) > 0 and miktar_col_fiziki in df.columns:
        first_row_val = safe_float(df.iloc[0][miktar_col_fiziki])
        if first_row_val > 0:
            root_qty = first_row_val

    # Ölçekleme Katsayısı
    scale_factor = hedef_adet / root_qty

    # 7. Hariç tutmaları uygula
    df_excluded = df.copy()
    if haric_tut:
        from recipe_automation.services.filters import apply_exclusions

        df_excluded, _, _ = apply_exclusions(df_excluded)

    # 8. ID/Stok filtre tipini belirle
    has_id = sira_col is not None
    has_rezerve = False
    for c in df_excluded.columns:
        if str(c).strip().lower() == str(settings.col_rezerve_miktar).strip().lower():
            has_rezerve = True
            break
    is_id_based = has_id and has_rezerve

    # TÜM_PARÇALAR seçilirse hedef grubu tüm operasyonlar yap
    if selected_group_name == "TÜM_PARÇALAR":
        all_ops = set()
        for col in op_cols:
            if col in df_excluded.columns:
                for val in df_excluded[col].dropna().unique():
                    all_ops.add(str(val).strip().upper())
        target_operations = all_ops

    # 9. ID veya Stok filtresini çalıştır
    if is_id_based:
        console.print("[green]ID-bazlı kırılım filtresi uygulanıyor...[/green]")
        from recipe_automation.services.filters import filter_id_based

        df_filtered, meta = filter_id_based(df_excluded, target_operations)
    else:
        console.print("[green]Stok-bazlı envanter filtresi uygulanıyor...[/green]")
        from recipe_automation.services.filters import filter_stock_based

        df_filtered, meta = filter_stock_based(df_excluded, target_operations)

    # 10. Ölçekleme işlemleri (Hem raw hem filtrelenmiş veri için)
    def scale_df(df_to_scale):
        df_sc = df_to_scale.copy().astype(object)
        target_quantities_keywords = [
            "miktar",
            "çarpılmış miktar",
            "carpilmis miktar",
            "rezerve edilecek miktar",
            "aktif rezerve edilen miktar",
            "kalan",
            "cikan",
        ]
        for col_name in df_sc.columns:
            nc = norm_col(str(col_name))
            if any(keyword in nc for keyword in target_quantities_keywords):
                df_sc[col_name] = df_sc[col_name].apply(
                    lambda val: safe_float(val) * scale_factor if pd.notna(val) else val
                )
        return df_sc

    df_raw_scaled = scale_df(df_excluded)
    df_filtered_scaled = scale_df(df_filtered)

    df_raw_scaled.insert(0, "KAYNAK DOSYA", base_file_name)
    df_filtered_scaled.insert(0, "KAYNAK DOSYA", base_file_name)

    # Geçici dosyayı kaydet
    temp_input_name = f"temp_scaled_input_{base_file_name}.xlsx"
    temp_input_path = os.path.join(input_dir, temp_input_name)
    df_filtered_scaled.to_excel(temp_input_path, index=False)

    # do_match_depo'yu çalıştır (tüm soruları soracak ve ek sayfaları oluşturacak)
    file_totals = {base_file_name: meta.get("orijinal_kalem_sayisi", 0)}

    for stock_basis in stock_bases:
        suffix_title = "FİZİKİ STOK" if stock_basis == "fiziki" else "KULLANILABİLİR STOK"
        console.print(
            f"\n[bold yellow]Filtreleme ve Eşleştirme Sistemi Çalıştırılıyor ({suffix_title} BAZINDA)...[/bold yellow]"
        )

        try:
            # Lazy import: main.py'deki circular import riskini önlemek için
            # do_match_depo yalnızca burada, ihtiyaç anında import ediliyor
            from recipe_automation.main import do_match_depo

            do_match_depo(
                temp_input_path,
                group=selected_group_name,
                file_totals=file_totals,
                raw_df=df_raw_scaled,
                input_path=file_path,
                stock_basis=stock_basis,
            )
        except Exception as e:
            console.print(f"[red]Eşleştirme aşamasında hata oluştu: {e}[/red]")
            continue

        temp_output_name = f"Filtered_{temp_input_name}"
        temp_output_path = os.path.join(input_dir, temp_output_name)

        if not os.path.exists(temp_output_path):
            console.print("[red]Hata: Eşleştirilmiş Excel dosyası oluşturulamadı.[/red]")
            continue

        # --- KAPASİTE ANALİZİ HESAPLAMALARI ---
        try:
            df_all = pd.read_excel(temp_output_path, sheet_name="Tüm Veriler")
        except Exception as e:
            console.print(f"[red]Tüm Veriler sayfası okunamadı: {e}[/red]")
            if os.path.exists(temp_output_path):
                os.remove(temp_output_path)
            continue

        fiziki_path = os.path.join(db_dir, "StokListesi.xlsx")

        stok_dict_kullanilabilir = {}
        stok_dict_fiziki = {}

        # 1. Öncelikle sürüklenen reçete Excel'indeki sütunlardan okumayı deneyelim
        kull_stok_col_in_df = None
        fiz_stok_col_in_df = None

        for c in df.columns:
            nc = norm_col(str(c))
            if "kullanilabilir" in nc and "stok" in nc:
                kull_stok_col_in_df = c
            elif nc == "stok" or (
                nc == "miktar"
                and "kritik" not in nc
                and "rezerve" not in nc
                and c != miktar_col_fiziki
                and c != miktar_col_kull
            ):
                if nc == "stok":
                    fiz_stok_col_in_df = c

        if kull_stok_col_in_df or fiz_stok_col_in_df:
            for _, row in df.iterrows():
                k = safe_kod(row[kod_col])
                if k and k != "nan":
                    if kull_stok_col_in_df:
                        stok_dict_kullanilabilir[k] = safe_float(row[kull_stok_col_in_df])
                    if fiz_stok_col_in_df:
                        stok_dict_fiziki[k] = safe_float(row[fiz_stok_col_in_df])

        # 2. Eğer veritabanı (StokListesi.xlsx) mevcutsa, verileri üzerine yaz/tamamla
        if os.path.exists(fiziki_path):
            try:
                with warnings.catch_warnings():
                    warnings.simplefilter("ignore")
                    df_stok = pd.read_excel(fiziki_path, header=0)

                kod_col_stok = None
                stok_col = None
                rezerve_col = None

                for c in df_stok.columns:
                    nc = norm_col(str(c))
                    if nc in ["malzeme kodu", "kod"]:
                        kod_col_stok = c
                    elif nc == "stok" or (nc == "miktar" and "kritik" not in nc):
                        stok_col = c
                    elif "rezerve" in nc:
                        rezerve_col = c

                if kod_col_stok and stok_col:
                    for _, row in df_stok.iterrows():
                        k = safe_kod(row[kod_col_stok])
                        if k and k != "nan":
                            fiziki_val = safe_float(row[stok_col])
                            rezerve_val = safe_float(row[rezerve_col]) if rezerve_col else 0.0
                            stok_dict_fiziki[k] = fiziki_val
                            stok_dict_kullanilabilir[k] = max(0.0, fiziki_val - rezerve_val)
            except Exception as e:
                console.print(f"[red]Stok listesi veritabanından okunurken hata: {e}[/red]")

        # Tüm Veriler'den miktar sütunlarını bul
        kod_col_all = None
        for c in df_all.columns:
            if str(c).strip().lower() == "kod":
                kod_col_all = c
                break

        col_kull = None
        for target in ["Rezerve Edilecek Miktar"]:
            for c in df_all.columns:
                if target.lower() in str(c).lower():
                    col_kull = c
                    break
            if col_kull:
                break

        col_fiziki = None
        for target in ["Çarpılmış Miktar", "Carpilmis Miktar", "Miktar"]:
            for c in df_all.columns:
                if target.lower() in str(c).lower():
                    col_fiziki = c
                    break
            if col_fiziki:
                break

        if not col_kull:
            col_kull = col_fiziki
        if not col_fiziki:
            col_fiziki = col_kull

        source_needs_kull = {}
        source_needs_fiz = {}

        if kod_col_all:
            for _, row in df_all.iterrows():
                k = safe_kod(row[kod_col_all])
                if k and k != "nan":
                    m_kull = safe_float(row[col_kull])
                    m_fiz = safe_float(row[col_fiziki])
                    unit_kull = m_kull / hedef_adet
                    unit_fiz = m_fiz / hedef_adet
                    source_needs_kull[k] = source_needs_kull.get(k, 0.0) + unit_kull
                    source_needs_fiz[k] = source_needs_fiz.get(k, 0.0) + unit_fiz

        # Analizleri yap
        analiz_kull = []
        max_k_kullanilabilir = 999999
        for kod, birim_ihtiyac in source_needs_kull.items():
            if birim_ihtiyac <= 0:
                continue
            stok_kull = stok_dict_kullanilabilir.get(kod, 0.0)
            max_k_this = math.floor(stok_kull / birim_ihtiyac)
            if max_k_this < max_k_kullanilabilir:
                max_k_kullanilabilir = max_k_this
            toplam_ihtiyac = birim_ihtiyac * hedef_adet
            eksik = max(0.0, toplam_ihtiyac - stok_kull)
            if eksik > 0:
                analiz_kull.append(
                    {
                        "Parça Kodu": kod,
                        "1 Adet İçin İhtiyaç": birim_ihtiyac,
                        f"{hedef_adet} Adet İçin İhtiyaç": toplam_ihtiyac,
                        "Kullanılabilir Stok": stok_kull,
                        "Net Eksik (Kullanılabilir)": eksik,
                    }
                )

        analiz_fiziki = []
        max_k_fiziki = 999999
        for kod, birim_ihtiyac in source_needs_fiz.items():
            if birim_ihtiyac <= 0:
                continue
            stok_fiz = stok_dict_fiziki.get(kod, 0.0)
            max_k_this = math.floor(stok_fiz / birim_ihtiyac)
            if max_k_this < max_k_fiziki:
                max_k_fiziki = max_k_this
            toplam_ihtiyac = birim_ihtiyac * hedef_adet
            eksik = max(0.0, toplam_ihtiyac - stok_fiz)
            if eksik > 0:
                analiz_fiziki.append(
                    {
                        "Parça Kodu": kod,
                        "1 Adet İçin İhtiyaç": birim_ihtiyac,
                        f"{hedef_adet} Adet İçin İhtiyaç": toplam_ihtiyac,
                        "Fiziki Stok": stok_fiz,
                        "Net Eksik (Fiziki)": eksik,
                    }
                )

        if max_k_kullanilabilir == 999999:
            max_k_kullanilabilir = 0
        if max_k_fiziki == 999999:
            max_k_fiziki = 0

        # Set bazlı yüzdelik hesaplama
        set_yuzdeleri = []
        for k in range(1, hedef_adet + 1):
            if stock_basis == "kullanilabilir":
                eksik_kull_say = sum(
                    1
                    for kod, birim in source_needs_kull.items()
                    if stok_dict_kullanilabilir.get(kod, 0.0) < (birim * k)
                )
                toplam_kull = len(source_needs_kull)
                y_kull = (
                    ((toplam_kull - eksik_kull_say) / toplam_kull) * 100 if toplam_kull > 0 else 0
                )
                set_yuzdeleri.append(
                    {
                        "Üretilecek Set": f"{k}. Set",
                        "Kull. Tamamlanma Oranı (%)": round(y_kull, 1),
                        "Kull. Eksik Kalemler": eksik_kull_say,
                    }
                )
            else:
                eksik_fiz_say = sum(
                    1
                    for kod, birim in source_needs_fiz.items()
                    if stok_dict_fiziki.get(kod, 0.0) < (birim * k)
                )
                toplam_fiz = len(source_needs_fiz)
                y_fiz = ((toplam_fiz - eksik_fiz_say) / toplam_fiz) * 100 if toplam_fiz > 0 else 0
                set_yuzdeleri.append(
                    {
                        "Üretilecek Set": f"{k}. Set",
                        "Fiziki Tamamlanma Oranı (%)": round(y_fiz, 1),
                        "Fiziki Eksik Kalemler": eksik_fiz_say,
                    }
                )

        console.print("\n[bold green]Analiz Tamamlandı![/bold green]")
        if stock_basis == "kullanilabilir":
            console.print(
                f"[bold yellow]=> Kullanılabilir Stoğa Göre Kapasite:[/bold yellow] [bold white]{max_k_kullanilabilir} Adet[/bold white]"
            )
        else:
            console.print(
                f"[bold cyan]=> Fiziki Stoğa Göre Kapasite:[/bold cyan] [bold white]{max_k_fiziki} Adet[/bold white]"
            )

        # --- GEÇİCİ RAPORU AÇIP KAPASİTE SAYFALARINI EKLEME ---
        wb = openpyxl.load_workbook(temp_output_path)

        # Kapasite sayfalarını en başa ekle
        ws_sim = wb.create_sheet("Kapasite Simülasyonu", 0)
        if stock_basis == "kullanilabilir":
            ws_kull = wb.create_sheet("Kullanılabilir Stok Eksikleri", 1)
            active_sheets = ["Kapasite Simülasyonu", "Kullanılabilir Stok Eksikleri"]
        else:
            ws_fiz = wb.create_sheet("Fiziki Stok Eksikleri", 1)
            active_sheets = ["Kapasite Simülasyonu", "Fiziki Stok Eksikleri"]

        # Verileri Yaz
        def write_data(ws, dict_list, default_msg=None):
            if not dict_list:
                if default_msg:
                    ws.append([default_msg])
                return
            headers = list(dict_list[0].keys())
            ws.append(headers)
            for item in dict_list:
                ws.append([item.get(h) for h in headers])

        write_data(ws_sim, set_yuzdeleri)
        if stock_basis == "kullanilabilir":
            write_data(ws_kull, analiz_kull, "Kullanılabilir stoğa göre eksik parça bulunmuyor.")
        else:
            write_data(ws_fiz, analiz_fiziki, "Fiziki stoğa göre eksik parça bulunmuyor.")

        # --- STİL VE HİZALAMA İŞLEMLERİ ---
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        rule_blue = DataBarRule(
            start_type="num",
            start_value=0,
            end_type="num",
            end_value=100,
            color="638EC6",
            showValue="None",
            minLength=None,
            maxLength=None,
        )
        rule_green = DataBarRule(
            start_type="num",
            start_value=0,
            end_type="num",
            end_value=100,
            color="5CB85C",
            showValue="None",
            minLength=None,
            maxLength=None,
        )

        for sh_name in active_sheets:
            ws = wb[sh_name]

            # Tüm hücrelere stil uygula
            for row in ws.iter_rows(
                min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column
            ):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if cell.row == 1:
                        cell.font = header_font
                        cell.fill = header_fill

            # Genişlikleri ayarla
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max(12, max_length + 4)
                ws.column_dimensions[column].width = adjusted_width

            # Simülasyon sayfasına veri çubuklarını ekle
            if sh_name == "Kapasite Simülasyonu":
                col_kull_oran = None
                col_fiz_oran = None
                for cell in ws[1]:
                    if "Kull. Tamamlanma" in str(cell.value):
                        col_kull_oran = cell.column_letter
                    elif "Fiziki Tamamlanma" in str(cell.value):
                        col_fiz_oran = cell.column_letter

                if col_kull_oran:
                    ws.conditional_formatting.add(
                        f"{col_kull_oran}2:{col_kull_oran}{ws.max_row}", rule_blue
                    )
                if col_fiz_oran:
                    ws.conditional_formatting.add(
                        f"{col_fiz_oran}2:{col_fiz_oran}{ws.max_row}", rule_green
                    )

        # Raporu Nihai İsimle Kaydet
        suffix = "FizikiStok" if stock_basis == "fiziki" else "KullStok"
        out_name = f"{base_file_name}_{hedef_adet}Adet_{selected_group_name}_{suffix}.xlsx"
        out_path = os.path.join(input_dir, out_name)

        try:
            wb.save(out_path)
            console.print(
                f"\n[bold green]Excel Raporu Başarıyla Oluşturuldu:[/bold green] [bold white]{out_name}[/bold white]\n"
            )
        except Exception as e:
            console.print(f"[red]Excel kaydedilirken hata oluştu: {e}[/red]")

        # Geçici çıktı dosyasını bu iterasyon için temizle
        if os.path.exists(temp_output_path):
            os.remove(temp_output_path)

    # Tüm iterasyonlar bittikten sonra geçici girdi dosyasını temizle
    if os.path.exists(temp_input_path):
        os.remove(temp_input_path)


if __name__ == "__main__":
    run_capacity_analysis()
