import json
import os
import re
import warnings

import openpyxl
import pandas as pd
import typer
from openpyxl.formatting.rule import CellIsRule, DataBarRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from rich.console import Console
from rich.panel import Panel
from rich.prompt import Confirm, Prompt
from rich.table import Table

from recipe_automation.core.config import settings
from recipe_automation.services.filters import filter_id_based, filter_stock_based
from recipe_automation.services.matcher import (
    append_carpimis_miktar,
    append_hammadde,
    match_with_depo,
)
from recipe_automation.services.scanner import extract_operations
from recipe_automation.services.sorter import load_priority_mapping, sort_dataframe
from recipe_automation.utils.excel_io import find_excel_files, read_excel_safe


def norm_col(text: str) -> str:
    if not text:
        return ""
    replacements = {
        "ı": "i",
        "İ": "i",
        "I": "i",
        "ş": "s",
        "Ş": "s",
        "ğ": "g",
        "Ğ": "g",
        "ü": "u",
        "Ü": "u",
        "ö": "o",
        "Ö": "o",
        "ç": "c",
        "Ç": "c",
    }
    t = str(text)
    for k, v in replacements.items():
        t = t.replace(k, v)
    return t.lower().replace(" ", "").replace("_", "")


warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

app = typer.Typer(help="Excel BOM / Reçete filtreleme aracı")
console = Console()


def load_group(group_name: str) -> set[str]:
    db_dir = os.path.join(os.getcwd(), settings.db_dir_name)
    json_path = os.path.join(db_dir, "operasyon_gruplari.json")
    if not os.path.exists(json_path):
        console.print(
            f"[red]Hata:[/red] {json_path} dosyası bulunamadı. Lütfen klasörde oluşturun."
        )
        raise typer.Exit(1)

    with open(json_path, encoding="utf-8") as f:
        data = json.load(f)

    if group_name == "TÜM_PARÇALAR":
        all_ops = set()
        for g_ops in data.values():
            for op in g_ops:
                all_ops.add(op.upper())
        return all_ops

    if group_name not in data:
        console.print(
            f"[red]Hata:[/red] '{group_name}' grubu JSON dosyasında bulunamadı. Mevcut gruplar: {list(data.keys())}"
        )
        raise typer.Exit(1)

    return {op.upper() for op in data[group_name]}


def print_report(file_name: str, meta: dict[str, str | int | list[str]]) -> None:
    table = Table(show_header=False, box=None)
    table.add_column("Key", style="cyan")
    table.add_column("Value", style="magenta")

    table.add_row("ID Sütunu", str(meta["kod_sutunu"]))
    table.add_row("Hedef Sütun", str(meta["hedef_sutun"]))

    op_cols = meta.get("op_cols", [])
    op_cols_str = ", ".join(op_cols) if isinstance(op_cols, list) and op_cols else "Yok"
    table.add_row("Taranan Operasyon Sütunları", op_cols_str)

    table.add_row("Başlangıç Satır Sayısı", str(meta["satir_ilk"]))
    table.add_row("Kalan Satır Sayısı", str(meta["satir_son"]))

    silinen_adet = meta.get("silinen_adet", 0)
    if silinen_adet > 0:
        table.add_row("JSON ile Silinen Satır Sayısı", str(silinen_adet))
        silinen_kodlar = meta.get("silinen_kodlar", [])
        if silinen_kodlar:
            kodlar_str = ", ".join(silinen_kodlar)
            table.add_row("Silinen Parça Kodları", f"[bold red]{kodlar_str}[/bold red]")

    panel = Panel(
        table,
        title=f"[bold green]Rapor: {file_name}[/bold green]",
        border_style="green",
        expand=False,
    )
    console.print(panel)


def aggregate_toplu_liste(df: pd.DataFrame) -> pd.DataFrame:
    """Birden fazla dosyayı birleştirirken mükerrer kodları toplayıp, Kaynak Dosya sütununa virgülle yazar."""
    kod_col = settings.col_sira_no_id
    if kod_col not in df.columns:
        for c in df.columns:
            if str(c).strip().lower() == "kod":
                kod_col = c
                break

    carp_col = None
    exact_targets = [
        "Çarpılmış Miktar",
        "Carpilmis Miktar",
        settings.col_rezerve_miktar,
        settings.col_kullanilabilir_stok,
    ]
    for target in exact_targets:
        if target in df.columns:
            carp_col = target
            break

    if not kod_col or not carp_col or "KAYNAK DOSYA" not in df.columns:
        from recipe_automation.services.filters import aggregate_duplicates

        return aggregate_duplicates(df)

    df = df.copy()
    df[kod_col] = df[kod_col].astype(str).str.strip().str.upper()
    df[carp_col] = pd.to_numeric(
        df[carp_col].astype(str).str.replace(",", "."), errors="coerce"
    ).fillna(0)

    res_list = []
    for k, group in df.groupby(kod_col, dropna=False):
        kaynak_list = []
        total_miktar = 0
        for _, row in group.iterrows():
            k_isim = str(row["KAYNAK DOSYA"]).strip()
            mikt = row[carp_col]
            total_miktar += mikt
            if mikt > 0:
                mikt_str = f"{mikt:g}"
                kaynak_list.append(f"{k_isim} ({mikt_str})")

        res_row = group.iloc[0].copy()
        if kaynak_list:
            res_row["KAYNAK DOSYA"] = ", ".join(kaynak_list)
        res_row[carp_col] = total_miktar
        res_list.append(res_row)

    df_grouped = pd.DataFrame(res_list)
    return df_grouped[df.columns]


@app.command()
def scan(path: str = typer.Argument(..., help="Excel dosya veya klasör yolu")) -> None:
    """Dosya veya klasördeki benzersiz operasyonları tarar ve EKRANA basar."""
    files = find_excel_files(path)
    if not files:
        console.print(f"[red]Hata:[/red] {path} yolunda Excel dosyası bulunamadı.")
        raise typer.Exit(1)

    db_dir = os.path.join(os.getcwd(), settings.db_dir_name)
    history_path = os.path.join(db_dir, "operasyon_gecmisi.txt")
    known_operations: set[str] = set()
    if os.path.exists(history_path):
        with open(history_path, encoding="utf-8") as f:
            for line in f:
                op = line.strip()
                if op:
                    known_operations.add(op.upper())

    current_operations: set[str] = set()
    for file in files:
        ops = set()
        cols = []
        for h in (0, 2):
            df = read_excel_safe(file, headers_to_try=(h,))
            if df is not None:
                ops, cols = extract_operations(df)
                if cols:
                    break

        console.print(
            f"[cyan]{os.path.basename(file)}[/cyan] -> Bulunan operasyon sütunları: {cols}"
        )
        if not cols:
            console.print(
                "[yellow]Uyarı:[/yellow] Bu dosyada 'Operasyon' kelimesi içeren sütun bulunamadı."
            )

        for op in ops:
            if str(op).strip() and str(op).lower() != "nan":
                current_operations.add(str(op).upper())

    if current_operations:
        new_operations = current_operations - known_operations
        all_operations = known_operations | current_operations

        console.print("\n[bold yellow]--- BU DOSYALARDA BULUNAN OPERASYONLAR ---[/bold yellow]")
        for op in sorted(list(current_operations)):
            console.print(f"  • {op}")
        console.print("------------------------------------\n")

        if new_operations:
            items_str = "\n".join(
                [f"[bold white] - {op}[/bold white]" for op in sorted(list(new_operations))]
            )
            panel = Panel(
                items_str,
                title="[bold red blink]!!! DİKKAT: YENİ OPERASYONLAR BULUNDU !!![/bold red blink]",
                border_style="red",
                expand=False,
            )
            console.print(panel)

            # JSON dosyasını oku
            db_dir = os.path.join(os.getcwd(), settings.db_dir_name)
            json_path = os.path.join(db_dir, "operasyon_gruplari.json")
            groups_data = {}
            if os.path.exists(json_path):
                with open(json_path, encoding="utf-8") as f:
                    groups_data = json.load(f)

            for new_op in sorted(list(new_operations)):
                console.print(
                    f"\n[bold cyan]Soru:[/bold cyan] [white]'{new_op}'[/white] operasyonunu hangi gruba dahil etmek istersin?"
                )
                group_keys = list(groups_data.keys())
                for i, g in enumerate(group_keys, 1):
                    console.print(f"  [{i}] {g}")
                console.print(f"  [{len(group_keys)+1}] + Yeni Bir Grup Oluştur")
                console.print(f"  [{len(group_keys)+2}] Hiçbir gruba ekleme (Atla)")

                choice_str = Prompt.ask(
                    "Seçiminiz", choices=[str(x) for x in range(1, len(group_keys) + 3)]
                )
                choice = int(choice_str)

                if choice <= len(group_keys):
                    selected_group = group_keys[choice - 1]
                    if new_op not in groups_data[selected_group]:
                        groups_data[selected_group].append(new_op)
                    console.print(
                        f"[green]✅ '{new_op}', '{selected_group}' grubuna eklendi.[/green]"
                    )
                elif choice == len(group_keys) + 1:
                    new_group_name = Prompt.ask("Yeni grubun adını girin (Örn: MONTAJ_GRUBU)")
                    new_group_name = new_group_name.strip().upper()
                    groups_data[new_group_name] = [new_op]
                    console.print(
                        f"[green]✅ '{new_group_name}' oluşturuldu ve operasyon eklendi.[/green]"
                    )
                else:
                    console.print("[yellow]⏭️ Atlandı. Sadece geçmişe kaydedildi.[/yellow]")

            # JSON'u kaydet
            with open(json_path, "w", encoding="utf-8") as f:
                json.dump(groups_data, f, indent=4, ensure_ascii=False)

            console.print(
                "\n[green]Tüm yeni operasyon atamaları JSON dosyasına kaydedildi![/green]\n"
            )

        with open(history_path, "w", encoding="utf-8") as f:
            for op in sorted(list(all_operations)):
                f.write(f"{op}\n")

        console.print(
            f"[green]✅ Başarılı![/green] Geçmiş dosyasında toplam {len(all_operations)} benzersiz operasyon kayıtlı."
        )
        console.print(f"Geçmiş dosyası: {history_path}")
    else:
        console.print("[yellow]Uyarı:[/yellow] Hiç operasyon bulunamadı.")


@app.command()
def auto_filter(
    path: str = typer.Argument(..., help="Excel dosya veya klasör yolu"),
    group: str = typer.Option("TIM", help="Filtrelenecek operasyon grubu (JSON'dan okur)"),
) -> None:
    """Verilen dosyayı inceleyerek ID mi yoksa Stok filtresi mi uygulanacağına otomatik karar verir."""
    files = find_excel_files(path)
    if not files:
        console.print(f"[red]Hata:[/red] {path} konumunda Excel dosyası bulunamadı.")
        raise typer.Exit(1)

    first_file = files[0]
    df = read_excel_safe(first_file, headers_to_try=(0, 2))

    if df is not None:
        has_id = settings.col_sira_no_id in df.columns
        has_rezerve = settings.col_rezerve_miktar in df.columns

        console.print(
            f"\n[bold magenta]🤖 Akıllı Analiz:[/bold magenta] '{os.path.basename(first_file)}' inceleniyor..."
        )
        if has_id and has_rezerve:
            console.print(
                "[green]✅ 'Sıra No' ve 'Rezerve Edilecek Miktar' sütunları bulundu. -> [bold]ID (Kırılım) Filtresi[/bold] başlatılıyor...[/green]"
            )
            filter_id(path, group)
        else:
            console.print(
                "[green]✅ Klasik kırılım sütunları bulunamadı. -> [bold]Stok (Envanter) Filtresi[/bold] başlatılıyor...[/green]"
            )
            filter_stock(path, group)
    else:
        console.print(f"[red]Hata:[/red] Dosya okunamadı: {first_file}")
        raise typer.Exit(1)


@app.command()
def filter_id(
    path: str = typer.Argument(..., help="Excel dosya veya klasör yolu"),
    group: str = typer.Option("TIM", help="Filtrelenecek operasyon grubu (JSON'dan okur)"),
) -> None:
    """ID (Sıra No) ve Rezerve Edilecek Miktar bazlı filtreleme yapar."""
    db_dir = os.path.join(os.getcwd(), settings.db_dir_name)
    priority_file = os.path.join(db_dir, "oncelik_sirasi.json")
    if not Confirm.ask(
        "[bold yellow]DİKKAT:[/bold yellow] Kaynak dosyalar için [white]'oncelik_sirasi.json'[/white] dosyasındaki öncelik sıralamasını güncellediniz mi?"
    ):
        console.print(
            "\n[red]İşlem iptal edildi. Lütfen öncelik sıralamanızı güncelleyip programı tekrar başlatın.[/red]"
        )
        if not os.path.exists(priority_file):
            os.makedirs(db_dir, exist_ok=True)
            with open(priority_file, "w", encoding="utf-8") as f:
                f.write('{\n  "ORNEK_KAYNAK_DOSYA_ADI": 1\n}')
        os.startfile(priority_file)
        raise typer.Exit()

    console.print("\n[bold cyan]=== ÖNCE YENİ OPERASYON TARAMASI YAPILIYOR ===[/bold cyan]")
    scan(path)
    db_dir = os.path.join(os.getcwd(), settings.db_dir_name)
    json_path = os.path.join(db_dir, "operasyon_gruplari.json")
    if os.path.exists(json_path):
        with open(json_path, encoding="utf-8") as f:
            groups_data = json.load(f)
        if groups_data:
            group_keys = list(groups_data.keys())
            console.print(
                "\n[bold cyan]Soru:[/bold cyan] Hangi grup için filtreleme yapmak istersiniz?"
            )
            for i, g in enumerate(group_keys, 1):
                console.print(f"  [{i}] {g}")
            choice_str = Prompt.ask(
                "Seçiminiz", choices=[str(x) for x in range(1, len(group_keys) + 1)]
            )
            group = group_keys[int(choice_str) - 1]

    console.print("\n[bold cyan]=== FİLTRELEME İŞLEMİ BAŞLATILIYOR ===[/bold cyan]")
    haric_tut = False
    if Confirm.ask(
        "[bold yellow]Soru:[/bold yellow] Merdane, boru, kaynak vb. parçalar (haric_tutulacak_parcalar.json listesindeki) yoksayılsın mı?",
        default=False,
    ):
        haric_tut = True

    target_ops = load_group(group)
    console.print(
        f"[bold blue]Kullanılan Grup:[/bold blue] {group} ({len(target_ops)} adet operasyon kuralı)"
    )

    files = find_excel_files(path)
    if not files:
        console.print("[red]Hata:[/red] Dosya bulunamadı.")
        raise typer.Exit(1)

    all_dfs = []
    all_raw_dfs = []
    file_totals = {}
    genel_silinen_kodlar = set()
    for file in files:
        success = False
        last_error = None
        for h in (0, 2):
            df = read_excel_safe(file, headers_to_try=(h,))
            if df is None:
                continue
            try:
                silinen_kodlar_dosya = []
                silinen_adet_dosya = 0
                if haric_tut:
                    from recipe_automation.services.filters import apply_exclusions

                    df, silinen_kodlar_dosya, silinen_adet_dosya = apply_exclusions(df)
                    if silinen_kodlar_dosya:
                        for kod in silinen_kodlar_dosya:
                            genel_silinen_kodlar.add(str(kod))

                res, meta = filter_id_based(df, target_ops)
                meta["silinen_kodlar"] = silinen_kodlar_dosya
                meta["silinen_adet"] = silinen_adet_dosya
                print_report(os.path.basename(file), meta)
                if not res.empty:
                    kaynak_metin = os.path.splitext(os.path.basename(file))[0]
                    carp_col = None
                    exact_targets = [
                        "Çarpılmış Miktar",
                        "Carpilmis Miktar",
                        settings.col_rezerve_miktar,
                        settings.col_kullanilabilir_stok,
                    ]
                    for target in exact_targets:
                        if target in res.columns:
                            carp_col = target
                            break

                    if carp_col:
                        console.print(
                            f"  [cyan]ℹ️ Miktar sütunu olarak '{carp_col}' tespit edildi ve kullanıldı.[/cyan]"
                        )
                        res.insert(0, "KAYNAK DOSYA", kaynak_metin)
                    else:
                        console.print(
                            "  [yellow]⚠️ Miktar sütunu BULUNAMADI! Varsayılan birleştirme uygulanıyor.[/yellow]"
                        )
                        res.insert(0, "KAYNAK DOSYA", kaynak_metin)
                    all_dfs.append(res)
                    df_raw_copy = df.copy()
                    df_raw_copy.insert(0, "KAYNAK DOSYA", kaynak_metin)
                    all_raw_dfs.append(df_raw_copy)
                    file_totals[kaynak_metin] = meta.get("orijinal_kalem_sayisi", 0)

                    if len(files) == 1:
                        out_path = os.path.join(
                            os.path.dirname(path), f"{group}_{os.path.basename(path)}"
                        )
                        res.to_excel(out_path, index=False)
                        console.print(f"[green]✅ Çıktı oluşturuldu:[/green] {out_path}\n")
                        console.print(
                            "\n[bold cyan]=== OTOMATİK DEPO EŞLEŞTİRME BAŞLATILIYOR ===[/bold cyan]"
                        )
                        do_match_depo(
                            out_path,
                            group=group,
                            file_totals=file_totals,
                            raw_df=df,
                            input_path=path,
                        )
                success = True
                break
            except ValueError as e:
                last_error = e
                continue

        if not success and last_error:
            console.print(f"[yellow]Atlanıyor {os.path.basename(file)}[/yellow]: {last_error}\n")

    if len(files) > 1 and all_dfs:
        combined = pd.concat(all_dfs, ignore_index=True)
        combined_agg = aggregate_toplu_liste(combined)
        folder_name = os.path.basename(os.path.normpath(path))
        out_path = os.path.join(path, f"{group}_{folder_name}.xlsx")
        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            combined_agg.to_excel(writer, sheet_name="TOPLU LİSTE", index=False)
            combined.to_excel(writer, sheet_name="RAW_DATA", index=False)
        console.print(f"[green]✅ Toplu dosya hazır:[/green] {out_path}")
        console.print("\n[bold cyan]=== OTOMATİK DEPO EŞLEŞTİRME BAŞLATILIYOR ===[/bold cyan]")
        combined_raw = pd.concat(all_raw_dfs, ignore_index=True) if all_raw_dfs else None
        do_match_depo(
            out_path, group=group, file_totals=file_totals, raw_df=combined_raw, input_path=path
        )

    if genel_silinen_kodlar:
        console.print("\n[bold red]=== JSON HARİÇ TUTMA ÖZETİ ===[/bold red]")
        console.print(
            f"[yellow]Tüm dosyalarda toplam {len(genel_silinen_kodlar)} adet benzersiz parça (ve alt kırılımları) silindi.[/yellow]"
        )
        console.print(
            f"[bold red]Silinen Tüm Kodlar:[/bold red] {', '.join(sorted(list(genel_silinen_kodlar)))}\n"
        )


@app.command()
def filter_stock(
    path: str = typer.Argument(..., help="Excel dosya veya klasör yolu"),
    group: str = typer.Option("TIM", help="Filtrelenecek operasyon grubu (JSON'dan okur)"),
) -> None:
    """Stok bazlı filtreleme yapar."""
    db_dir = os.path.join(os.getcwd(), settings.db_dir_name)
    priority_file = os.path.join(db_dir, "oncelik_sirasi.json")
    if not Confirm.ask(
        "[bold yellow]DİKKAT:[/bold yellow] Kaynak dosyalar için [white]'oncelik_sirasi.json'[/white] dosyasındaki öncelik sıralamasını güncellediniz mi?"
    ):
        console.print(
            "\n[red]İşlem iptal edildi. Lütfen öncelik sıralamanızı güncelleyip programı tekrar başlatın.[/red]"
        )
        if not os.path.exists(priority_file):
            os.makedirs(db_dir, exist_ok=True)
            with open(priority_file, "w", encoding="utf-8") as f:
                f.write('{\n  "ORNEK_KAYNAK_DOSYA_ADI": 1\n}')
        os.startfile(priority_file)
        raise typer.Exit()

    console.print("\n[bold cyan]=== ÖNCE YENİ OPERASYON TARAMASI YAPILIYOR ===[/bold cyan]")
    scan(path)

    db_dir = os.path.join(os.getcwd(), settings.db_dir_name)
    json_path = os.path.join(db_dir, "operasyon_gruplari.json")
    if os.path.exists(json_path):
        with open(json_path, encoding="utf-8") as f:
            groups_data = json.load(f)
        if groups_data:
            group_keys = list(groups_data.keys())
            console.print(
                "\n[bold cyan]Soru:[/bold cyan] Hangi grup için filtreleme yapmak istersiniz?"
            )
            for i, g in enumerate(group_keys, 1):
                console.print(f"  [{i}] {g}")
            choice_str = Prompt.ask(
                "Seçiminiz", choices=[str(x) for x in range(1, len(group_keys) + 1)]
            )
            group = group_keys[int(choice_str) - 1]

    console.print("\n[bold cyan]=== FİLTRELEME İŞLEMİ BAŞLATILIYOR ===[/bold cyan]")
    haric_tut = False
    if Confirm.ask(
        "[bold yellow]Soru:[/bold yellow] Merdane, boru, kaynak vb. parçalar (haric_tutulacak_parcalar.json listesindeki) yoksayılsın mı?",
        default=False,
    ):
        haric_tut = True

    target_ops = load_group(group)
    console.print(
        f"[bold blue]Kullanılan Grup:[/bold blue] {group} ({len(target_ops)} adet operasyon kuralı)"
    )

    files = find_excel_files(path)
    if not files:
        console.print("[red]Hata:[/red] Dosya bulunamadı.")
        raise typer.Exit(1)

    all_dfs = []
    all_raw_dfs = []
    file_totals = {}
    genel_silinen_kodlar = set()
    for file in files:
        success = False
        last_error = None
        for h in (2, 0):  # Reçete genelde 2'den başlar
            df = read_excel_safe(file, headers_to_try=(h,))
            if df is None:
                continue
            try:
                silinen_kodlar_dosya = []
                silinen_adet_dosya = 0
                if haric_tut:
                    from recipe_automation.services.filters import apply_exclusions

                    df, silinen_kodlar_dosya, silinen_adet_dosya = apply_exclusions(df)
                    if silinen_kodlar_dosya:
                        for kod in silinen_kodlar_dosya:
                            genel_silinen_kodlar.add(str(kod))

                res, meta = filter_stock_based(df, target_ops)
                meta["silinen_kodlar"] = silinen_kodlar_dosya
                meta["silinen_adet"] = silinen_adet_dosya
                print_report(os.path.basename(file), meta)
                if not res.empty:
                    kaynak_metin = os.path.splitext(os.path.basename(file))[0]
                    carp_col = None
                    exact_targets = [
                        "Çarpılmış Miktar",
                        "Carpilmis Miktar",
                        settings.col_rezerve_miktar,
                        settings.col_kullanilabilir_stok,
                    ]
                    for target in exact_targets:
                        if target in res.columns:
                            carp_col = target
                            break

                    if carp_col:
                        console.print(
                            f"  [cyan]ℹ️ Miktar sütunu olarak '{carp_col}' tespit edildi ve kullanıldı.[/cyan]"
                        )
                        res.insert(0, "KAYNAK DOSYA", kaynak_metin)
                    else:
                        console.print(
                            "  [yellow]⚠️ Miktar sütunu BULUNAMADI! Varsayılan birleştirme uygulanıyor.[/yellow]"
                        )
                        res.insert(0, "KAYNAK DOSYA", kaynak_metin)
                    all_dfs.append(res)
                    df_raw_copy = df.copy()
                    df_raw_copy.insert(0, "KAYNAK DOSYA", kaynak_metin)
                    all_raw_dfs.append(df_raw_copy)
                    file_totals[kaynak_metin] = meta.get("orijinal_kalem_sayisi", 0)

                    if len(files) == 1:
                        out_path = os.path.join(
                            os.path.dirname(path), f"{group}_{os.path.basename(path)}"
                        )
                        res.to_excel(out_path, index=False)
                        console.print(f"[green]✅ Çıktı oluşturuldu:[/green] {out_path}\n")
                        console.print(
                            "\n[bold cyan]=== OTOMATİK DEPO EŞLEŞTİRME BAŞLATILIYOR ===[/bold cyan]"
                        )
                        do_match_depo(
                            out_path,
                            group=group,
                            file_totals=file_totals,
                            raw_df=df,
                            input_path=path,
                        )
                success = True
                break
            except ValueError as e:
                last_error = e
                continue

        if not success and last_error:
            console.print(f"[yellow]Atlanıyor {os.path.basename(file)}[/yellow]: {last_error}\n")

    if len(files) > 1 and all_dfs:
        combined = pd.concat(all_dfs, ignore_index=True)
        combined_agg = aggregate_toplu_liste(combined)
        folder_name = os.path.basename(os.path.normpath(path))
        out_path = os.path.join(path, f"{group}_{folder_name}.xlsx")
        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            combined_agg.to_excel(writer, sheet_name="TOPLU LİSTE", index=False)
            combined.to_excel(writer, sheet_name="RAW_DATA", index=False)
        console.print(f"[green]✅ Toplu dosya hazır:[/green] {out_path}")
        console.print("\n[bold cyan]=== OTOMATİK DEPO EŞLEŞTİRME BAŞLATILIYOR ===[/bold cyan]")
        combined_raw = pd.concat(all_raw_dfs, ignore_index=True) if all_raw_dfs else None
        do_match_depo(
            out_path, group=group, file_totals=file_totals, raw_df=combined_raw, input_path=path
        )

    if genel_silinen_kodlar:
        console.print("\n[bold red]=== JSON HARİÇ TUTMA ÖZETİ ===[/bold red]")
        console.print(
            f"[yellow]Tüm dosyalarda toplam {len(genel_silinen_kodlar)} adet benzersiz parça (ve alt kırılımları) silindi.[/yellow]"
        )
        console.print(
            f"[bold red]Silinen Tüm Kodlar:[/bold red] {', '.join(sorted(list(genel_silinen_kodlar)))}\n"
        )


@app.command()
def match_depo(
    path: str = typer.Argument(..., help="Filtrelenmiş Excel dosyasının yolu"),
    group: str = typer.Option("TIM", help="Hangi grup için eşleştirme yapıldığı"),
) -> None:
    """Filtrelenmiş listeyi Ana Depo ve Hammadde Deposu ile eşleştirir."""
    do_match_depo(path, group=group, file_totals=None, raw_df=None, input_path=None)


def do_match_depo(
    path: str,
    group: str = "TIM",
    file_totals: dict = None,
    raw_df: pd.DataFrame = None,
    input_path: str = None,
    stock_basis: str = "kullanilabilir",
) -> None:
    """Core logic for match_depo, isolated to support dict arguments which Typer rejects."""
    parent_stats = {}
    stok_dict_kullanilabilir = {}
    if not os.path.exists(path) or not os.path.isfile(path):
        console.print(f"[red]Hata:[/red] Geçerli bir dosya sürüklemediniz: {path}")
        raise typer.Exit(1)

    db_dir = os.path.join(os.getcwd(), settings.db_dir_name)
    if not os.path.exists(db_dir):
        os.makedirs(db_dir, exist_ok=True)
        console.print(
            f"[yellow]Uyarı:[/yellow] '{settings.db_dir_name}' klasörü bulunamadığı için otomatik oluşturuldu."
        )
        console.print(
            "[red]Hata:[/red] Lütfen 'TumRotaBilgileri.xlsx' ve 'ReceteTumRotaListe.xlsx' dosyalarını bu klasöre koyup tekrar deneyin."
        )
        raise typer.Exit(1)

    depo_path = os.path.join(db_dir, "TumRotaBilgileri.xlsx")
    hammadde_path = os.path.join(db_dir, "ReceteTumRotaListe.xlsx")

    if not os.path.exists(depo_path):
        console.print(f"[red]Hata:[/red] Ana depo bulunamadı: {depo_path}")
        raise typer.Exit(1)
    if not os.path.exists(hammadde_path):
        console.print(f"[red]Hata:[/red] Hammadde deposu bulunamadı: {hammadde_path}")
        raise typer.Exit(1)

    console.print("\n[bold blue]1. Veritabanı:[/bold blue] TumRotaBilgileri.xlsx")
    console.print("[bold blue]2. Veritabanı:[/bold blue] ReceteTumRotaListe.xlsx")

    filtered_df = read_excel_safe(path, headers_to_try=(0, 2))

    # RAW veriyi okumayı dene (Üretim Takip sayfası için)
    try:
        ham_takip_df = pd.read_excel(path, sheet_name="RAW_DATA")
    except Exception:
        ham_takip_df = filtered_df.copy()

    depo_df = read_excel_safe(depo_path, headers_to_try=(0, 2))
    hammadde_df = read_excel_safe(hammadde_path, headers_to_try=(0, 2))

    if filtered_df is None:
        console.print(f"[red]Hata:[/red] Filtrelenmiş dosya okunamadı: {path}")
        raise typer.Exit(1)
    if depo_df is None:
        console.print(f"[red]Hata:[/red] Ana depo dosyası okunamadı: {depo_path}")
        raise typer.Exit(1)
    if hammadde_df is None:
        console.print(f"[red]Hata:[/red] Hammadde dosyası okunamadı: {hammadde_path}")
        raise typer.Exit(1)

    try:
        matched_df = match_with_depo(filtered_df, depo_df)

        # Aktif/Pasif (1-0, Evet-Hayır) Temizliği Soruları
        target_check_cols = ["receteaktifmi", "anarotami", "rotaaktifmi", "istasyonanakayitmi"]
        found_cols = []
        for c in matched_df.columns:
            if norm_col(c) in target_check_cols and c not in found_cols:
                found_cols.append(c)

        if found_cols:
            console.print(
                "\n[bold cyan]Soru:[/bold cyan] Depo excelinde aktif/pasif durumunu belirten bazı kritik sütunlar bulundu."
            )
            for col in found_cols:
                if col not in matched_df.columns:
                    continue
                if matched_df.empty:
                    break

                ans = Confirm.ask(
                    f"[white]'{col}'[/white] sütununda değeri '0' veya 'Hayır' (Pasif) olan satırları LİSTEDEN SİLELİM Mİ?",
                    default=True,
                )
                if ans:
                    before_len = len(matched_df)

                    s_data = matched_df[col]
                    if isinstance(s_data, pd.DataFrame):
                        s_data = s_data.iloc[:, 0]

                    # 0, '0', 'hayır', 'hayir', 'false' varyasyonlarını temizle
                    mask = (
                        s_data.astype(str)
                        .str.strip()
                        .str.lower()
                        .apply(
                            lambda x: x.replace(".0", "")
                            not in ["0", "hayir", "hayır", "false", "pasif", "nan", "none"]
                        )
                        .astype(bool)
                    )
                    matched_df = matched_df[mask].copy()
                    console.print(
                        f"  [green]✅ Temizlendi![/green] (Bu işlemle {before_len - len(matched_df)} satır daha ayıklandı)"
                    )

        import json

        # Tanımlanan Gruplar: JSON'dan oku
        istasyon_file = os.path.join(db_dir, "istasyon_gruplari.json")
        if not os.path.exists(istasyon_file):
            default_mapping = {
                "TIM": {
                    "SN50": ["TOS2"],
                    "SN71": ["TOS3"],
                    "CY": ["NEXT110Y", "TS4000Y-1", "TS4000Y-2"],
                    "NEX110": ["NEXT110"],
                    "QUASER": ["QUASER-1", "QUASER-2", "QUASER-3"],
                    "PHOEBUS": ["PHOEBUS"],
                    "ARION2142": ["ARİON2142KÖPRÜ", "ARION2142KOPRU"],
                    "ARION2000S": ["ARİONGSM2000", "ARIONGSM2000"],
                    "MATKAP&KILAVUZ": ["MATKAP+KILAVUZ", "KILAVUZ", "MATKAP", "MATKAP+KILAVUZ"],
                    "TMX2000S": ["TMX2000S"],
                    "PLANYA": ["PLANYA"],
                    "SA32B": ["NEXTURNSA32B"],
                    "HAMMADDE": ["TESTERE", "EBATLAMA"],
                    "3D PRINT": ["3DYAZICI"],
                },
                "BUKUM": {},
                "KAYNAK": {},
                "LAZER KESIM": {},
            }
            with open(istasyon_file, "w", encoding="utf-8") as f:
                json.dump(default_mapping, f, indent=4, ensure_ascii=False)

        with open(istasyon_file, encoding="utf-8") as f:
            tum_istasyon_gruplari = json.load(f)

        operasyon_file = os.path.join(db_dir, "operasyon_gruplari.json")
        operasyon_gruplari = {}
        if os.path.exists(operasyon_file):
            with open(operasyon_file, encoding="utf-8") as f:
                operasyon_gruplari = json.load(f)

        # Kapasite ayarlarını yükle
        kapasite_file = os.path.join(db_dir, "kapasite_ayarlari.json")
        kapasite_ayarlari = {}
        if os.path.exists(kapasite_file):
            try:
                with open(kapasite_file, encoding="utf-8") as f:
                    kapasite_ayarlari = json.load(f)
            except Exception as e:
                console.print(f"[red]Kapasite ayarları dosyası okunurken hata: {e}[/red]")

        def get_station_capacity_settings(name):
            varsayilan_saat = kapasite_ayarlari.get("varsayilan_gunluk_saat", 9)
            istasyon_ayarlar = kapasite_ayarlari.get("istasyonlar", {})

            # Tam eşleşme dene
            if name in istasyon_ayarlar:
                cfg = istasyon_ayarlar[name]
                return cfg.get("gunluk_saat", varsayilan_saat), cfg.get("makine_sayisi")

            # Normalize ederek dene
            name_norm = name.strip().upper().replace(" ", "")
            for key, cfg in istasyon_ayarlar.items():
                if key.strip().upper().replace(" ", "") == name_norm:
                    return cfg.get("gunluk_saat", varsayilan_saat), cfg.get("makine_sayisi")

            return varsayilan_saat, None

        if group == "TÜM_PARÇALAR":
            station_mapping = {}
            for g_name, g_map in tum_istasyon_gruplari.items():
                for sheet_name, keywords in g_map.items():
                    if sheet_name not in station_mapping:
                        station_mapping[sheet_name] = []
                    for kw in keywords:
                        if kw not in station_mapping[sheet_name]:
                            station_mapping[sheet_name].append(kw)
        else:
            station_mapping = tum_istasyon_gruplari.get(group, {})

        # Hammadde Eklemesi (Left Join)
        # Sadece istasyon_gruplari.json içinde "HAMMADDE" dizisi olan gruplar için hammadde bilgisini ekle
        if "HAMMADDE" in station_mapping:
            matched_df = append_hammadde(matched_df, hammadde_df)

        # Çarpılmış Miktar Eklemesi
        matched_df = append_carpimis_miktar(matched_df, filtered_df)

        # --- YENİ EKLENTİ: KAPASİTE ANALİZİNDE STOK DÜŞME (NET GEREKSİNİM) ---
        is_capacity_mode = os.path.basename(path).startswith("temp_scaled_input_")
        if is_capacity_mode:
            stok_path = os.path.join(db_dir, "StokListesi.xlsx")

            # 1. Öncelikle girdi reçete dosyasından eldeki kullanılabilir stokları oku (varsa)
            orig_path = input_path if input_path else path
            if os.path.exists(orig_path):
                try:
                    with warnings.catch_warnings():
                        warnings.simplefilter("ignore")
                        df_orig = pd.read_excel(orig_path, header=None)

                    h_idx = -1
                    for i_row, row_val in df_orig.iterrows():
                        row_str = " ".join([str(x).lower() for x in row_val.values])
                        if "kod" in row_str and ("malzeme" in row_str or "ad" in row_str):
                            h_idx = i_row
                            break

                    if h_idx != -1:
                        df_orig_clean = df_orig.iloc[h_idx + 1 :].copy()
                        df_orig_clean.columns = df_orig.iloc[h_idx]
                        df_orig_clean = df_orig_clean.dropna(how="all")

                        kod_col_orig = None
                        kull_stok_col_in_df = None

                        for c in df_orig_clean.columns:
                            nc = norm_col(str(c))
                            if nc in ["kod", "malzeme kodu"]:
                                kod_col_orig = c
                            if stock_basis == "fiziki":
                                if "fiziki" in nc and "stok" in nc:
                                    kull_stok_col_in_df = c
                                    break
                        if stock_basis == "fiziki" and not kull_stok_col_in_df:
                            for c in df_orig_clean.columns:
                                nc = norm_col(str(c))
                                if (
                                    "stok" in nc
                                    and "kullanilabilir" not in nc
                                    and "rezerve" not in nc
                                ):
                                    kull_stok_col_in_df = c
                                    break
                        if not kull_stok_col_in_df:
                            for c in df_orig_clean.columns:
                                nc = norm_col(str(c))
                                if "kullanilabilir" in nc and "stok" in nc:
                                    kull_stok_col_in_df = c
                                    break

                        if kod_col_orig and kull_stok_col_in_df:
                            for _, row in df_orig_clean.iterrows():
                                k = str(row[kod_col_orig]).strip().upper()
                                if k.endswith(".0"):
                                    k = k[:-2]
                                if k and k != "NAN" and k != "":
                                    if pd.notna(row[kull_stok_col_in_df]):
                                        try:
                                            stok_dict_kullanilabilir[k] = float(
                                                str(row[kull_stok_col_in_df]).replace(",", ".")
                                            )
                                        except:
                                            pass
                except Exception as e:
                    console.print(
                        f"  [yellow]Uyarı:[/yellow] Orijinal dosyadan stok okuma hatası: {e}"
                    )

            # 2. StokListesi.xlsx veritabanından oku ve üzerine yaz/tamamla
            if os.path.exists(stok_path):
                try:
                    with warnings.catch_warnings():
                        warnings.simplefilter("ignore")
                        df_stok = pd.read_excel(stok_path, header=0)

                    kod_col_stok = None
                    stok_col = None
                    rezerve_col = None

                    for c in df_stok.columns:
                        nc = norm_col(str(c))
                        if nc in ["malzeme kodu", "kod"]:
                            kod_col_stok = c
                        elif nc == "stok" or (nc == "miktar" and "kritik" not in nc):
                            stok_col = c
                        elif "rezerve" in nc:
                            rezerve_col = c

                    if kod_col_stok and stok_col:
                        for _, row in df_stok.iterrows():
                            k = str(row[kod_col_stok]).strip().upper()
                            if k.endswith(".0"):
                                k = k[:-2]
                            if k and k != "NAN" and k != "":
                                try:
                                    fiziki_val = float(str(row[stok_col]).replace(",", "."))
                                except:
                                    fiziki_val = 0.0
                                try:
                                    rezerve_val = (
                                        float(str(row[rezerve_col]).replace(",", "."))
                                        if rezerve_col and pd.notna(row[rezerve_col])
                                        else 0.0
                                    )
                                except:
                                    rezerve_val = 0.0

                                if stock_basis == "fiziki":
                                    stok_dict_kullanilabilir[k] = fiziki_val
                                else:
                                    stok_dict_kullanilabilir[k] = max(0.0, fiziki_val - rezerve_val)
                except Exception as e:
                    console.print(
                        f"  [yellow]Uyarı:[/yellow] StokListesi veritabanından stok okuma hatası: {e}"
                    )

            # 3. matched_df'deki 'Üretilecek Miktar'ı net eksik ile güncelle
            if "Üretilecek Miktar" in matched_df.columns:

                def get_net_qty(row):
                    k = str(row.get(settings.col_depo_kod, "")).strip().upper()
                    if k.endswith(".0"):
                        k = k[:-2]
                    gross_qty = 0.0
                    try:
                        gross_qty = float(str(row["Üretilecek Miktar"]).replace(",", "."))
                    except:
                        pass
                    stok_kull = stok_dict_kullanilabilir.get(k, 0.0)
                    return max(0.0, gross_qty - stok_kull)

                matched_df["Üretilecek Miktar"] = matched_df.apply(get_net_qty, axis=1)

        # --- YENİ EKLENTİ: TOPLAM SÜRE ---
        toplam_sure_col = "Toplam Süre"
        if "Birim İşlem Süresi" in matched_df.columns:
            if toplam_sure_col in matched_df.columns:
                matched_df = matched_df.drop(columns=[toplam_sure_col])
            cols = list(matched_df.columns)
            insert_idx = cols.index("Birim İşlem Süresi") + 1
            matched_df.insert(insert_idx, toplam_sure_col, 0)

        console.print(
            "\n[green]✅ İkinci Veritabanından Hammaddeler ve Filtrelenmiş Dosyadan Miktarlar başarıyla eklendi![/green]"
        )
        console.print("  [cyan]⏳ İstasyon verileri işleniyor ve gruplara ayrılıyor...[/cyan]")

        kaynak_dosya_adi = os.path.basename(path)
        out_path = os.path.join(os.path.dirname(path), f"Filtered_{kaynak_dosya_adi}")

        # Diğer sayfalardaki kısa format: filtered_df'deki KAYNAK DOSYA sütunundan al (örn: "2241")
        # Yoksa dosya adından türet: "TIM_2241.xlsx" → "2241"
        kaynak_metin_kisa = kaynak_dosya_adi
        if "KAYNAK DOSYA" in filtered_df.columns:
            vals = filtered_df["KAYNAK DOSYA"].dropna().unique()
            if len(vals) > 0:
                kaynak_metin_kisa = str(vals[0])
        else:
            # TIM_2241.xlsx → 2241
            ad = os.path.splitext(kaynak_dosya_adi)[0]  # TIM_2241
            for prefix in ["Filtered_TIM_", "Filtered_", "TIM_"]:
                if ad.startswith(prefix):
                    ad = ad[len(prefix) :]
                    break
            kaynak_metin_kisa = ad

        # --- YENİ EKLENTİ: KAYNAK DOSYA SÜTUNU ---
        if "KAYNAK DOSYA" not in matched_df.columns:
            matched_df.insert(0, "KAYNAK DOSYA", kaynak_dosya_adi)

        # İstasyon İsimlerini Temizleme Fonksiyonu
        def normalize_station_name(val):
            if pd.isna(val):
                return ""
            return str(val).strip().upper().replace(" ", "")

        # Tanımlanan Gruplar (station_mapping) yukarıda (hammadde eklemesi öncesinde) JSON'dan okundu.

        machine_counts = {}
        for original_name, machine_list in station_mapping.items():
            safe_name = re.sub(r"[\\/*?:\[\]]", "-", str(original_name))[:31]
            machine_counts[safe_name] = max(1, len(machine_list))

        # İç kullanım için normalize et
        station_mapping_normalized = {}
        for sheet, keywords in station_mapping.items():
            station_mapping_normalized[sheet] = [normalize_station_name(k) for k in keywords]

        # İstasyon Sütununu Bul
        istasyon_col = None
        for c in matched_df.columns:
            if "istasyonu" in norm_col(c) or "isistasyonu" in norm_col(c):
                istasyon_col = c
                break

        # Operasyon Adı Sütununu Bul (Hammadde için)
        operasyon_col = None
        for c in matched_df.columns:
            if "operasyon" in norm_col(c) and "adi" in norm_col(c):
                operasyon_col = c
                break

        # Operasyon Sıra No Sütununu Bul (Hammadde sıra no kısıtı için)
        operasyon_sira_col = None
        for c in matched_df.columns:
            if (
                "operasyon" in norm_col(c)
                and ("sira" in norm_col(c) or "no" in norm_col(c))
                and "adi" not in norm_col(c)
            ):
                operasyon_sira_col = c
                break

        # --- YENİ EKLENTİ: ROTASIZLAR SAYFASI İÇİN FİLTRELEME ---
        console.print("  [cyan]⏳ Rotasız parçalar tespit ediliyor ve ayrıştırılıyor...[/cyan]")
        source_df = raw_df if raw_df is not None else filtered_df
        op_cols = [
            c for c in source_df.columns if settings.col_operasyon_keyword.lower() in str(c).lower()
        ]
        df_rotasiz = pd.DataFrame()
        if op_cols:
            first_op_col = op_cols[0]
            first_op_val = source_df[first_op_col].astype(str).str.strip()
            is_empty_mask = (
                source_df[first_op_col].isna()
                | (first_op_val == "")
                | (first_op_val.str.lower() == "nan")
                | (first_op_val == "0")
            )
            df_rotasiz = source_df[is_empty_mask].copy()
        else:
            df_rotasiz = source_df.copy()

        sheet_dfs = {}
        unassigned_stations = set()

        if istasyon_col:
            for idx, row in matched_df.iterrows():
                st_val = row[istasyon_col]
                norm_st = normalize_station_name(st_val)

                op_val = row[operasyon_col] if operasyon_col else ""
                norm_op = normalize_station_name(op_val)

                # Operasyon sıra nosunu al ve temizle
                op_seq = ""
                if operasyon_sira_col:
                    seq_val = row[operasyon_sira_col]
                    if pd.notna(seq_val):
                        seq_str = str(seq_val).strip()
                        if seq_str.endswith(".0"):
                            seq_str = seq_str[:-2]
                        op_seq = seq_str

                assigned_sheets = []

                for sheet_name, keywords in station_mapping_normalized.items():
                    if sheet_name == "HAMMADDE":
                        matched_ham = False
                        for kw in keywords:
                            if ":" in kw:
                                parts = kw.split(":")
                                target_op = parts[0]
                                target_seq = parts[1]
                                if norm_op == target_op and op_seq == target_seq:
                                    matched_ham = True
                                    break
                            else:
                                if norm_op == kw:
                                    matched_ham = True
                                    break
                        if matched_ham:
                            assigned_sheets.append(sheet_name)
                    else:
                        if norm_st in keywords:
                            assigned_sheets.append(sheet_name)

                if assigned_sheets:
                    for sh in assigned_sheets:
                        if sh not in sheet_dfs:
                            sheet_dfs[sh] = []
                        sheet_dfs[sh].append(row)
                else:
                    if str(st_val).strip() != "" and str(st_val).lower() != "nan":
                        unassigned_stations.add(str(st_val).strip())

            for sheet_name, rows in sheet_dfs.items():
                sheet_dfs[sheet_name] = pd.DataFrame(rows)

        # --- YENİ EKLENTİ: ROTASIZLAR SAYFASINI EKLE ---
        if not df_rotasiz.empty:
            if "KAYNAK DOSYA" not in df_rotasiz.columns:
                df_rotasiz.insert(0, "KAYNAK DOSYA", kaynak_dosya_adi)
            sheet_dfs["Rotasızlar"] = df_rotasiz

        # --- YENİ EKLENTİ: MONTAJ İZLEME SAYFASI ---
        if raw_df is not None:
            console.print(
                "  [cyan]⏳ Montaj izleme ilişkileri ve hiyerarşi ağacı analiz ediliyor...[/cyan]"
            )
            # "MONTAJ OTOMASYON" ve "FINAL MONTAJ" kelimelerini içeren operasyonların listesi
            montaj_keywords = []
            op_to_group = {}
            for ana_grup, alt_gruplar in operasyon_gruplari.items():
                if "MONTAJ" in str(ana_grup).upper():
                    for op in alt_gruplar:
                        op_str = str(op).upper().strip().replace(" ", "")
                        montaj_keywords.append(op_str)
                        op_to_group[op_str] = str(ana_grup).upper().strip()

            # --- DEBUG ---
            op_cols = [
                c
                for c in raw_df.columns
                if settings.col_operasyon_keyword.lower() in str(c).lower()
            ]
            op_col = op_cols[0] if op_cols else None

            # Sira No sutununu bul - once tam eslesme, sonra norm_col ile
            sira_col = None
            if settings.col_sira_no_id in raw_df.columns:
                sira_col = settings.col_sira_no_id
            else:
                for c in raw_df.columns:
                    if norm_col(c) in ["sirano", "sirano", "sira"]:
                        sira_col = c
                        break

            if op_col and sira_col:
                raw_df["norm_op"] = (
                    raw_df[op_col].astype(str).str.strip().str.upper().str.replace(" ", "")
                )
                montaj_parents_df = raw_df[raw_df["norm_op"].isin(montaj_keywords)].copy()

                # TIM grubuna ait operasyon kelimeleri
                target_ops_for_raw = load_group(group)

                # raw_df'de Kod sütununu bul (parça numarası)
                raw_kod_col = None
                for c in raw_df.columns:
                    if str(c).strip().lower() == "kod":
                        raw_kod_col = c
                        break

                # filtered_df'den Sıra No → Gereken Miktar haritası
                miktar_col = None
                for c in [
                    settings.col_rezerve_miktar,
                    settings.col_kullanilabilir_stok,
                    "Çarpılmış Miktar",
                    "Carpilmis Miktar",
                ]:
                    if c in filtered_df.columns:
                        miktar_col = c
                        break

                # Kalan miktarları Kod veya Sıra No bazında kümülatif toplayalım
                kod_to_remaining_qty = {}
                filt_id_col = (
                    raw_kod_col
                    if (raw_kod_col and raw_kod_col in filtered_df.columns)
                    else sira_col
                )

                if miktar_col and filt_id_col in filtered_df.columns:
                    for _, r in filtered_df.iterrows():
                        k_val = str(r.get(filt_id_col, "")).strip().upper()
                        if k_val and k_val != "nan":
                            try:
                                qty_val = float(str(r.get(miktar_col, 0)).replace(",", "."))
                            except ValueError:
                                qty_val = 0.0
                            kod_to_remaining_qty[k_val] = (
                                kod_to_remaining_qty.get(k_val, 0.0) + qty_val
                            )

                # Eşleşen veritabanı üzerinden TIM operasyonuna sahip olan tüm parça kodlarının setini alalım
                tim_codes_in_db = set(
                    depo_df[
                        depo_df["Operasyon Adı"]
                        .astype(str)
                        .str.strip()
                        .str.upper()
                        .isin(target_ops_for_raw)
                    ]["Kod"]
                    .dropna()
                    .astype(str)
                    .str.strip()
                    .str.upper()
                )

                # raw_df'de TIM operasyonuna sahip veya veritabanında TIM rotası olan satırları bul
                def _is_tim(row):
                    # 1. Kod veya Sıra No'ya göre veritabanında TIM rotası var mı kontrol et
                    row_kod = str(row.get(raw_kod_col or "Kod", "")).strip().upper()
                    if row_kod in tim_codes_in_db:
                        return True
                    # 2. Reçetedeki operasyon sütunlarına göre kontrol et (fallback)
                    for oc in [
                        c
                        for c in raw_df.columns
                        if settings.col_operasyon_keyword.lower() in str(c).lower()
                    ]:
                        val = str(row.get(oc, "")).strip().upper()
                        if val and val in target_ops_for_raw:
                            return True
                    return False

                tim_mask_raw = raw_df.apply(_is_tim, axis=1)
                tim_rows_raw = raw_df[tim_mask_raw].copy()
                parent_stats = {}

                if not montaj_parents_df.empty and not tim_rows_raw.empty:
                    # Tüm alt parçaların toplam reçete ihtiyacını hesapla
                    # NOT: matched_df değil raw_df kullanılıyor — çünkü rotasız parçalar veya
                    # başka sayfaya düşen parçaların da mevcut sipariş ihtiyaçları stoktan düşülmeli.
                    # Kullanıcı hangi filtreleme grubuyla çalıştıysa o siparişin tüm parçaları kapsanmalı.
                    total_required_qty = {}
                    # Öncelik: "Miktar" → sipariş için o parçadan reçetede kaç adet gerektiğini gösterir.
                    # "Rezerve Edilecek Miktar" 0 olabilir (parça zaten rezerve edilmiş olabilir),
                    # o yüzden bunu değil gerçek reçete miktarını kullanıyoruz.
                    raw_qty_col = None
                    for col_name in [
                        "Miktar",
                        "Çarpılmış Miktar",
                        "Carpilmis Miktar",
                        settings.col_rezerve_miktar,
                        "Üretilecek Miktar",
                    ]:
                        if col_name in raw_df.columns:
                            raw_qty_col = col_name
                            break
                    raw_id_col = raw_kod_col or settings.col_depo_kod
                    for _, row in raw_df.iterrows():
                        k = str(row.get(raw_id_col, "")).strip().upper()
                        if k.endswith(".0"):
                            k = k[:-2]
                        qty = 0.0
                        if raw_qty_col:
                            try:
                                qty = float(str(row.get(raw_qty_col, 0)).replace(",", "."))
                            except:
                                pass
                        if k and k != "NAN" and k != "":
                            total_required_qty[k] = total_required_qty.get(k, 0.0) + qty

                    montaj_izleme_dict = {}

                    # Önce tüm TIM satırlarını montaj parent'larıyla eşleştirip listeleyelim
                    matched_tim_relations = []
                    for idx, row in tim_rows_raw.iterrows():
                        sira_no = str(row.get(sira_col, "")).strip()
                        if not sira_no or sira_no == "nan":
                            continue

                        # 1.002.014.005 → prefix'ler: ["1", "1.002", "1.002.014"]
                        parts = sira_no.split(".")
                        prefixes = [".".join(parts[:i]) for i in range(1, len(parts))]

                        # Bu prefix'lerden biri montaj_parents_df içindeyse eşleşme var
                        if "KAYNAK DOSYA" in montaj_parents_df.columns and "KAYNAK DOSYA" in row:
                            row_kaynak = row.get("KAYNAK DOSYA")
                            parents = montaj_parents_df[
                                (montaj_parents_df[sira_col].astype(str).isin(prefixes))
                                & (montaj_parents_df["KAYNAK DOSYA"] == row_kaynak)
                            ]
                        else:
                            parents = montaj_parents_df[
                                montaj_parents_df[sira_col].astype(str).isin(prefixes)
                            ]
                        if not parents.empty:
                            # En yakın (hiyerarşide en derindeki) üst montajı bul
                            # Bunun için Sıra No içindeki nokta sayısına (seviyesine) göre en büyüğünü seçiyoruz
                            p_row = parents.loc[
                                parents[sira_col]
                                .astype(str)
                                .apply(lambda x: len(x.split(".")))
                                .idxmax()
                            ]

                            p_op = str(p_row.get("norm_op", ""))
                            m_grup = op_to_group.get(p_op, "MONTAJ")
                            target_sheet = f"{m_grup} İZLEME"

                            parca_adi = row.get("Malzeme", row.get("Malzeme Adı", ""))
                            parca_kodu = row.get(raw_kod_col, sira_no) if raw_kod_col else sira_no

                            # Orijinal miktar
                            try:
                                orig_qty = float(
                                    str(
                                        row.get(miktar_col, row.get(settings.col_rezerve_miktar, 1))
                                    ).replace(",", ".")
                                )
                            except ValueError:
                                orig_qty = 1.0

                            # Tasarım miktarı (stoktakiler dahil toplam miktar)
                            design_qty_col = None
                            for qc in ["Miktar", "Adet", "Qty", "Quantity", miktar_col]:
                                if qc in row:
                                    design_qty_col = qc
                                    break
                            try:
                                design_qty = float(
                                    str(row.get(design_qty_col or miktar_col, 1)).replace(",", ".")
                                )
                            except ValueError:
                                design_qty = orig_qty

                            matched_tim_relations.append(
                                {
                                    "row_ref": row,
                                    "target_sheet": target_sheet,
                                    "p_row": p_row,
                                    "parca_kodu": str(parca_kodu).strip().upper(),
                                    "parca_adi": parca_adi,
                                    "orig_qty": orig_qty,
                                    "design_qty": design_qty,
                                    "sira_no": sira_no,
                                }
                            )

                    # Eşleşen ilişkilerin istatistiklerini topla (Toplam ve Eksik Çeşit Takibi İçin)
                    for rel in matched_tim_relations:
                        kaynak = str(rel["row_ref"].get("KAYNAK DOSYA", kaynak_metin_kisa)).strip()
                        parent_kod = str(
                            rel["p_row"].get("Kod", rel["p_row"].get(sira_col, ""))
                        ).strip()
                        child_kod = str(rel["parca_kodu"]).strip().upper()

                        key = (kaynak, parent_kod)
                        if key not in parent_stats:
                            parent_stats[key] = {"total": set(), "missing": set(), "total_qty": 0.0}
                        parent_stats[key]["total"].add(child_kod)
                        parent_stats[key]["total_qty"] += rel.get("design_qty", rel["orig_qty"])

                        key_val = (
                            rel["parca_kodu"] if filt_id_col == raw_kod_col else rel["sira_no"]
                        )
                        remaining_qty = kod_to_remaining_qty.get(key_val, 0.0)
                        if remaining_qty > 0:
                            parent_stats[key]["missing"].add(child_kod)

                    # Şimdi bu ilişkileri parça_kodu (veya sira_no) bazında gruplayıp kalan miktarları dağıtalım
                    grouped_relations = {}
                    for rel in matched_tim_relations:
                        key_val = (
                            rel["parca_kodu"] if filt_id_col == raw_kod_col else rel["sira_no"]
                        )
                        if key_val not in grouped_relations:
                            grouped_relations[key_val] = []
                        grouped_relations[key_val].append(rel)

                    for key_val, rel_list in grouped_relations.items():
                        # Kalan miktar
                        remaining_qty = kod_to_remaining_qty.get(key_val, 0.0)
                        if remaining_qty <= 0:
                            # Eğer bu parça filtrelenmiş listede kalmadıysa (stok yetti veya elendi) montaj listesine EKLEME!
                            continue

                        # Kalan miktarı orijinal miktarlara göre dağıt
                        temp_qty = remaining_qty
                        for rel in rel_list:
                            allocated = min(rel["orig_qty"], temp_qty)
                            temp_qty -= allocated
                            rel["allocated_qty"] = allocated

                        # Kalan küsurat varsa sonuncu ilişkiye ekle
                        if temp_qty > 0 and len(rel_list) > 0:
                            rel_list[-1]["allocated_qty"] += temp_qty

                        # Şimdi izleme dict'ine ekle
                        for rel in rel_list:
                            alloc_qty = rel.get("allocated_qty", 0.0)
                            if alloc_qty > 0:
                                t_sheet = rel["target_sheet"]
                                if t_sheet not in montaj_izleme_dict:
                                    montaj_izleme_dict[t_sheet] = []

                                montaj_izleme_dict[t_sheet].append(
                                    {
                                        "Kaynak Dosya": rel["row_ref"].get(
                                            "KAYNAK DOSYA", kaynak_metin_kisa
                                        ),
                                        "Üst Montaj Kodu": rel["p_row"].get(
                                            "Kod", rel["p_row"].get(sira_col, "")
                                        ),
                                        "Üst Montaj Adı": rel["p_row"].get(
                                            "Malzeme", rel["p_row"].get("Malzeme Adı", "")
                                        ),
                                        "Alt Parça Kodu": rel["parca_kodu"],
                                        "Alt Parça Adı": rel["parca_adi"],
                                        "Gereken Miktar": alloc_qty,
                                        "Üretilen Miktar": 0,
                                        "Tamamlanma Oranı (%)": 0,
                                    }
                                )

                    if montaj_izleme_dict:
                        priority_mapping = load_priority_mapping(db_dir)
                        from recipe_automation.services.sorter import calculate_row_priority

                        for sheet_name, rows in montaj_izleme_dict.items():
                            df_izleme = pd.DataFrame(rows)
                            # Aynı child aynı parent için birden fazla eklendiyse topla
                            df_izleme = df_izleme.groupby(
                                [
                                    "Kaynak Dosya",
                                    "Üst Montaj Kodu",
                                    "Üst Montaj Adı",
                                    "Alt Parça Kodu",
                                    "Alt Parça Adı",
                                ],
                                as_index=False,
                            ).agg(
                                {
                                    "Gereken Miktar": "sum",
                                    "Üretilen Miktar": "sum",
                                    "Tamamlanma Oranı (%)": "max",
                                }
                            )
                            if not df_izleme.empty:
                                temp_prio = df_izleme["Kaynak Dosya"].apply(
                                    lambda x: calculate_row_priority(x, priority_mapping)
                                )
                                df_izleme = (
                                    df_izleme.assign(tmp_prio=temp_prio)
                                    .sort_values(
                                        by=[
                                            "tmp_prio",
                                            "Kaynak Dosya",
                                            "Üst Montaj Kodu",
                                            "Alt Parça Kodu",
                                        ]
                                    )
                                    .drop(columns=["tmp_prio"])
                                )
                            sheet_dfs[sheet_name] = df_izleme

        # --- YENİ EKLENTİ: HAMMADDE SİPARİŞİ SAYFASI ---
        if "HAMMADDE" in sheet_dfs:
            df_ham = sheet_dfs["HAMMADDE"]
            hk_col = settings.col_hammadde_kod
            h_col = settings.col_hammadde_isim
            thm_col = "Toplam Hammadde Miktarı"

            if hk_col in df_ham.columns and thm_col in df_ham.columns:
                # Toplamlar artık Excel formülü ile atılacak. Sadece benzersiz kodları alıyoruz.
                df_ham_copy = df_ham.copy()
                df_ham_copy[thm_col] = 0

                agg_dict = {thm_col: "first"}
                if h_col in df_ham_copy.columns:
                    agg_dict[h_col] = "first"

                df_siparis = df_ham_copy.groupby(hk_col, as_index=False).agg(agg_dict)

                # Sütun sıralamasını düzenle
                if h_col in df_siparis.columns:
                    df_siparis = df_siparis[[hk_col, h_col, thm_col]]
                else:
                    df_siparis = df_siparis[[hk_col, thm_col]]

                sheet_dfs["HAMMADDE SİPARİŞ"] = df_siparis

        # --- YENİ EKLENTİ: ÜRETİM LİSTESİ SAYFASI ---
        uretim_cols = []
        if "KAYNAK DOSYA" in matched_df.columns:
            uretim_cols.append("KAYNAK DOSYA")
        uretim_cols.append(settings.col_depo_kod)

        if "Ad" in matched_df.columns:
            uretim_cols.append("Ad")
        elif "Malzeme Adı" in matched_df.columns:
            uretim_cols.append("Malzeme Adı")
        elif "Malzeme" in matched_df.columns:
            uretim_cols.append("Malzeme")

        if settings.col_hammadde_kod in matched_df.columns:
            uretim_cols.append(settings.col_hammadde_kod)
        if settings.col_hammadde_isim in matched_df.columns:
            uretim_cols.append(settings.col_hammadde_isim)

        if "Çarpılmış Miktar" in matched_df.columns:
            uretim_cols.append("Çarpılmış Miktar")
        if "Üretilecek Miktar" in matched_df.columns:
            uretim_cols.append("Üretilecek Miktar")

        df_uretim = matched_df[uretim_cols].drop_duplicates(subset=[settings.col_depo_kod]).copy()

        # --- YENİ EKLENTİ: ÜRETİM TAKİP SAYFASI (Ayrı Kaynak Dosyalar) ---
        console.print("  [cyan]⏳ Üretim takip sayfası hazırlanıyor...[/cyan]")
        # ham_takip_df yukarida RAW_DATA sekmesinden veya fallback olarak filtered_df'den okundu
        filtered_kod_sutunu = None
        for c in ["Kod", settings.col_sira_no_id]:
            if c in ham_takip_df.columns:
                filtered_kod_sutunu = c
                break

        if filtered_kod_sutunu and settings.col_depo_kod in depo_df.columns:
            depo_kodlar = set(
                depo_df[settings.col_depo_kod].dropna().astype(str).str.strip().str.upper()
            )
            filt_kodlar = (
                ham_takip_df[filtered_kod_sutunu].dropna().astype(str).str.strip().str.upper()
            )
            gecerli_maske = filt_kodlar.isin(depo_kodlar)
            takip_gecerli = ham_takip_df[gecerli_maske].copy()

            carp_col = None
            for c in [
                "Çarpılmış Miktar",
                "Carpilmis Miktar",
                settings.col_rezerve_miktar,
                settings.col_kullanilabilir_stok,
            ]:
                if c in takip_gecerli.columns:
                    carp_col = c
                    break

            cols_to_keep = ["KAYNAK DOSYA"] if "KAYNAK DOSYA" in takip_gecerli.columns else []
            cols_to_keep.append(filtered_kod_sutunu)
            if carp_col:
                cols_to_keep.append(carp_col)

            takip_df = takip_gecerli[cols_to_keep].copy()

            rename_dict = {filtered_kod_sutunu: "Kod"}
            if carp_col:
                rename_dict[carp_col] = "Üretilecek Miktar"
            takip_df = takip_df.rename(columns=rename_dict)

            takip_df["Üretilen Miktar"] = 0
            takip_df["Kalan Miktar"] = 0
            takip_df["Tamamlanma (%)"] = 0.0

            if "KAYNAK DOSYA" in takip_df.columns:
                takip_df["KAYNAK DOSYA"] = (
                    takip_df["KAYNAK DOSYA"]
                    .astype(str)
                    .str.strip()
                    .apply(lambda x: x[:-2] if x.endswith(".0") else x)
                )
                takip_df = takip_df.sort_values(by=["KAYNAK DOSYA", "Kod"])

            if is_capacity_mode and "Üretilecek Miktar" in takip_df.columns:

                def get_net_takip(row):
                    k = str(row.get("Kod", "")).strip().upper()
                    if k.endswith(".0"):
                        k = k[:-2]
                    gross_qty = 0.0
                    try:
                        gross_qty = float(str(row["Üretilecek Miktar"]).replace(",", "."))
                    except:
                        pass
                    stok_kull = stok_dict_kullanilabilir.get(k, 0.0)
                    return max(0.0, gross_qty - stok_kull)

                takip_df["Üretilecek Miktar"] = takip_df.apply(get_net_takip, axis=1)

            sheet_dfs["Üretim Takip"] = takip_df

        # --- YENİ EKLENTİ: KAYNAK DOSYA ÖNCELİK SIRALAMASI ---
        priority_mapping = load_priority_mapping(db_dir)
        df_uretim = sort_dataframe(df_uretim, priority_mapping)
        matched_df = sort_dataframe(matched_df, priority_mapping)

        for s_name, s_df in sheet_dfs.items():
            sheet_dfs[s_name] = sort_dataframe(s_df, priority_mapping)

        # --- YENİ EKLENTİ: METİN OLARAK SAKLANAN SAYILARI GERÇEK SAYIYA ÇEVİRME ---
        def convert_numbers_to_real(df):
            def to_real_number(val):
                if pd.isna(val):
                    return val
                if isinstance(val, (int, float)):
                    return val
                val_str = str(val).strip()
                if val_str.lstrip("-").isdigit():
                    return int(val_str)
                try:
                    return float(val_str)
                except ValueError:
                    return val

            for col in df.columns:
                if col != "KAYNAK DOSYA":  # Kaynak dosya isimleri metin kalmalı (örn "2254 (2)")
                    df[col] = df[col].apply(to_real_number)
            return df

        df_uretim = convert_numbers_to_real(df_uretim)
        matched_df = convert_numbers_to_real(matched_df)
        for s_name in list(sheet_dfs.keys()):
            sheet_dfs[s_name] = convert_numbers_to_real(sheet_dfs[s_name])
            if (
                s_name
                not in [
                    "ÜRETİM LİSTESİ",
                    "HAMMADDE SİPARİŞ",
                    "Tüm Veriler",
                    "HAMMADDE",
                    "Üretim Takip",
                    "Rotasızlar",
                ]
                and "ZLEME" not in s_name.upper()
            ):
                sheet_dfs[s_name]["Setup Yükü (%)"] = None
                sheet_dfs[s_name]["Önerilen Verimli Adet"] = None
                sheet_dfs[s_name]["Güncel Setup Yükü (%)"] = None

        # ExcelWriter ile Çoklu Sheet Yazdırma
        console.print("  [cyan]⏳ Excel sayfaları diske yazılıyor (Pandas)...[/cyan]")
        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            # 1. Ana Kumanda Sayfası: ÜRETİM LİSTESİ
            df_uretim.to_excel(writer, sheet_name="ÜRETİM LİSTESİ", index=False)

            # 2. Ana Sayfa (Tüm Veriler)
            matched_df.to_excel(writer, sheet_name="Tüm Veriler", index=False)

            # 2. HAMMADDE Sayfası (Eğer varsa)
            if "HAMMADDE" in sheet_dfs:
                sheet_dfs["HAMMADDE"].to_excel(writer, sheet_name="HAMMADDE", index=False)

            # 3. HAMMADDE SİPARİŞİ Sayfası (Eğer varsa)
            if "HAMMADDE SİPARİŞ" in sheet_dfs:
                sheet_dfs["HAMMADDE SİPARİŞ"].to_excel(
                    writer, sheet_name="HAMMADDE SİPARİŞ", index=False
                )

            # 4. Üretim Takip Sayfası
            if "Üretim Takip" in sheet_dfs:
                sheet_dfs["Üretim Takip"].to_excel(writer, sheet_name="Üretim Takip", index=False)

            # 4.2 ROTASIZLAR Sayfası
            if "Rotasızlar" in sheet_dfs:
                sheet_dfs["Rotasızlar"].to_excel(writer, sheet_name="Rotasızlar", index=False)

            # 4.1 İZLEME Sayfaları (Dinamik - operasyon_gruplari.json'dan gelen MONTAJ grupları)
            izleme_sheets = [s for s in sheet_dfs.keys() if "ZLEME" in s.upper()]

            def sort_izleme_key(name):
                n_upper = name.upper()
                if "MONTAJ OTOMASYON" in n_upper:
                    return 0
                elif "FINAL MONTAJ" in n_upper or "FİNAL MONTAJ" in n_upper:
                    return 1
                return 2

            izleme_sheets_sorted = sorted(izleme_sheets, key=sort_izleme_key)
            for s_name in izleme_sheets_sorted:
                df_izleme = sheet_dfs[s_name]
                safe_name = re.sub(r"[\\/*?:\[\]]", "-", str(s_name))[:31]
                df_izleme.to_excel(writer, sheet_name=safe_name, index=False)

            # 5. Geri Kalan Diğer İstasyon Sayfaları
            def natural_keys(text):
                import re

                return [int(c) if c.isdigit() else c.lower() for c in re.split(r"(\d+)", str(text))]

            sorted_sheet_names = sorted(sheet_dfs.keys(), key=natural_keys)

            for sheet_name in sorted_sheet_names:
                df_sheet = sheet_dfs[sheet_name]
                if (
                    sheet_name not in ["HAMMADDE", "HAMMADDE SİPARİŞ", "Üretim Takip", "Rotasızlar"]
                    and "ZLEME" not in sheet_name.upper()
                ):
                    safe_name = re.sub(r"[\\/*?:\[\]]", "-", str(sheet_name))[:31]
                    df_sheet.to_excel(writer, sheet_name=safe_name, index=False)

        # --- İSTENEN SÜTUNLARI GİZLEME VE GENİŞLİK AYARLAMA ---
        target_hide_cols = [
            "birimagirlik",
            "receteonaydurumu",
            "receteaktifmi",
            "rotaadi",
            "anarotami",
            "rotaaktifmi",
            "rotaistikameti",
            "operasyonadi",
            "rotaoperasyononaydurumu",
            "istasyonanakayitmi",
            "fasonmu",
            "rezerveedilecekmiktar",
            "onceliksirasi",
        ]

        # Kullanıcının talebi: JSON'da hammadde tanımlıysa (TIM vb.), Malzeme Adı gizlensin.
        # Tanımlı değilse (Montaj Otomasyon vb.), Malzeme Adı gösterilsin.
        if "HAMMADDE" in station_mapping:
            target_hide_cols.append("malzemeadi")

        # Sadece istasyon sayfalarında (Tüm Veriler, HAMMADDE ve SİPARİŞ harici) gizlenecek ekstra sütunlar
        station_extra_hide = [
            "hammaddemiktar",
            "toplamhammaddemiktari",
            "isistasyonu",
            "operasyonsirano",
        ]

        try:
            max_df_len = max(len(matched_df), len(df_uretim) if "df_uretim" in locals() else 0)
            for df_val in sheet_dfs.values():
                if isinstance(df_val, pd.DataFrame):
                    max_df_len = max(max_df_len, len(df_val))
            global_max_row = max_df_len + 2000
            console.print(
                f"  [cyan]ℹ️ Dinamik Excel Limiti Ayarlandı: Maksimum Sayfa Satır Sayısı = {max_df_len}, Belirlenen Limit = {global_max_row}[/cyan]"
            )
            console.print(
                "  [cyan]⏳ Excel dosyası openpyxl ile yükleniyor (Stiller ve formüller uygulanacak)...[/cyan]"
            )
            wb = openpyxl.load_workbook(out_path)
            wb.parent_stats = parent_stats
            gizlenen_adet = 0

            # Dosyadaki TÜM sheetler için aynı ayarları uygula
            console.print(
                "  [cyan]⏳ Formüller enjekte ediliyor, hücre stilleri ve koşullu biçimlendirmeler uygulanıyor...[/cyan]"
            )
            for ws_name in wb.sheetnames:
                ws = wb[ws_name]
                is_station_sheet = (
                    ws_name
                    not in [
                        "Tüm Veriler",
                        "HAMMADDE",
                        "HAMMADDE SİPARİŞ",
                        "Üretim Takip",
                        "ÜRETİM LİSTESİ",
                        "Rotasızlar",
                    ]
                    and "ZLEME" not in ws_name.upper()
                )

                for col in ws.columns:
                    col_letter = col[0].column_letter
                    header_val = col[0].value

                    hide_this_col = False
                    if ws_name == "Üretim Takip" and col_letter in [
                        "A",
                        "B",
                        "C",
                        "D",
                        "E",
                        "F",
                        "G",
                        "H",
                    ]:
                        hide_this_col = True
                    elif header_val:
                        n_col = norm_col(header_val)
                        if n_col in target_hide_cols:
                            hide_this_col = True
                        elif is_station_sheet and n_col in station_extra_hide:
                            hide_this_col = True

                    if hide_this_col:
                        ws.column_dimensions[col_letter].hidden = True
                        if ws_name == "Tüm Veriler":
                            gizlenen_adet += 1
                    else:
                        max_length = 0
                        col_cells = list(col)
                        sample_cells = col_cells[:100]
                        if len(col_cells) > 100:
                            sample_cells.extend(col_cells[-10:])
                        for cell in sample_cells:
                            val = cell.value
                            if val is not None:
                                try:
                                    val_len = len(str(val))
                                    if val_len > max_length:
                                        max_length = val_len
                                except:
                                    pass
                        adjusted_width = min(max_length + 2, 45)
                        ws.column_dimensions[col_letter].width = adjusted_width

                # --- YENİ EKLENTİ: EXCEL FORMÜLLERİNİ ENJEKTE ETME ---
                # Öncelikle sütun harflerini bul
                uretilecek_col_letter = None
                ham_miktar_col_letter = None
                toplam_col_letter = None
                hk_col_letter = None
                dk_col_letter = None

                hazirlik_col_letter = None
                birim_col_letter = None
                toplam_sure_col_letter = None

                setup_yuku_col_letter = None
                onerilen_adet_col_letter = None
                guncel_setup_col_letter = None

                for col in ws.columns:
                    val = str(col[0].value).strip() if col[0].value else ""
                    if val == "Üretilecek Miktar":
                        uretilecek_col_letter = col[0].column_letter
                    elif val == "Hammadde Miktar":
                        ham_miktar_col_letter = col[0].column_letter
                    elif val == "Toplam Hammadde Miktarı":
                        toplam_col_letter = col[0].column_letter
                    elif val == settings.col_hammadde_kod:
                        hk_col_letter = col[0].column_letter
                    elif val == settings.col_depo_kod:
                        dk_col_letter = col[0].column_letter
                    elif val == "Hazırlık Süresi":
                        hazirlik_col_letter = col[0].column_letter
                    elif val == "Birim İşlem Süresi" or val == "Birim İşlem":
                        birim_col_letter = col[0].column_letter
                    elif val == "Toplam Süre":
                        toplam_sure_col_letter = col[0].column_letter
                    elif val == "Setup Yükü (%)":
                        setup_yuku_col_letter = col[0].column_letter
                    elif val == "Önerilen Verimli Adet":
                        onerilen_adet_col_letter = col[0].column_letter
                    elif val == "Güncel Setup Yükü (%)":
                        guncel_setup_col_letter = col[0].column_letter

                # Eğer HAMMADDE sayfasıysa harflerini global olarak aklımızda tutalım (SUMIF için) ve sayfayı gizleyelim
                if ws_name == "HAMMADDE":
                    wb.hammadde_kod_harfi = hk_col_letter
                    wb.hammadde_toplam_harfi = toplam_col_letter
                    ws.sheet_state = "hidden"

                # Formülleri Yazdır
                if ws_name in ["HAMMADDE SİPARİŞ", "HAMMADDE SİPARİŞİ"]:
                    if (
                        hk_col_letter
                        and toplam_col_letter
                        and hasattr(wb, "hammadde_kod_harfi")
                        and hasattr(wb, "hammadde_toplam_harfi")
                    ):
                        # SUMIF Formülü (Örn: =SUMIF(HAMMADDE!B:B, A2, HAMMADDE!E:E))
                        for row_idx in range(2, ws.max_row + 1):
                            cell_kod = f"{hk_col_letter}{row_idx}"
                            formula = f"=SUMIF(HAMMADDE!{wb.hammadde_kod_harfi}:{wb.hammadde_kod_harfi}, {cell_kod}, HAMMADDE!{wb.hammadde_toplam_harfi}:{wb.hammadde_toplam_harfi})"
                            ws[f"{toplam_col_letter}{row_idx}"] = formula
                elif ws_name not in ["ÜRETİM LİSTESİ", "Üretim Takip", "Rotasızlar"]:
                    # Alt sayfalardaki Üretilecek Miktarı Ana Sayfaya (ÜRETİM LİSTESİ) DÜŞEYARA (VLOOKUP) ile bağla
                    if uretilecek_col_letter and dk_col_letter:
                        # VLOOKUP arama aralığı Excel'de daima aranan kelimenin sütunundan başlamalıdır!
                        # Örneğin Kod B sütunundaysa arama A:Z değil B:Z olmalı.
                        actual_uretim_cols = df_uretim.columns.tolist()
                        idx_of_kod_in_uretim = actual_uretim_cols.index(settings.col_depo_kod)
                        start_letter = chr(65 + idx_of_kod_in_uretim)

                        if "Üretilecek Miktar" in actual_uretim_cols:
                            lookup_idx = (
                                actual_uretim_cols.index("Üretilecek Miktar")
                                - idx_of_kod_in_uretim
                                + 1
                            )
                        else:
                            lookup_idx = len(actual_uretim_cols) - idx_of_kod_in_uretim

                        for row_idx in range(2, ws.max_row + 1):
                            cell_kod = f"{dk_col_letter}{row_idx}"
                            vlookup_formula = f"=VLOOKUP({cell_kod}, 'ÜRETİM LİSTESİ'!{start_letter}:Z, {lookup_idx}, FALSE)"
                            ws[f"{uretilecek_col_letter}{row_idx}"] = vlookup_formula

                    if uretilecek_col_letter and ham_miktar_col_letter and toplam_col_letter:
                        # ÇARPIM Formülü (Örn: =D2*E2)
                        for row_idx in range(2, ws.max_row + 1):
                            formula = f"={uretilecek_col_letter}{row_idx}*{ham_miktar_col_letter}{row_idx}"
                            ws[f"{toplam_col_letter}{row_idx}"] = formula

                    if (
                        uretilecek_col_letter
                        and hazirlik_col_letter
                        and birim_col_letter
                        and toplam_sure_col_letter
                    ):
                        # TOPLAM SÜRE Formülü (Örn: =U2+(V2*C2))
                        for row_idx in range(2, ws.max_row + 1):
                            formula = f"={hazirlik_col_letter}{row_idx}+({birim_col_letter}{row_idx}*{uretilecek_col_letter}{row_idx})"
                            ws[f"{toplam_sure_col_letter}{row_idx}"] = formula

                    # YENİ EKLENTİ: Setup ve Verimlilik Formülleri
                    if uretilecek_col_letter and hazirlik_col_letter and birim_col_letter:
                        if setup_yuku_col_letter:
                            for row_idx in range(2, ws.max_row + 1):
                                ws[f"{setup_yuku_col_letter}{row_idx}"] = (
                                    f"={hazirlik_col_letter}{row_idx}/({uretilecek_col_letter}{row_idx}*{birim_col_letter}{row_idx})"
                                )
                                ws[f"{setup_yuku_col_letter}{row_idx}"].number_format = "0%"

                        if onerilen_adet_col_letter and setup_yuku_col_letter:
                            for row_idx in range(2, ws.max_row + 1):
                                ws[f"{onerilen_adet_col_letter}{row_idx}"] = (
                                    f"=IF({setup_yuku_col_letter}{row_idx}<=0.15, {uretilecek_col_letter}{row_idx}, ROUNDUP({hazirlik_col_letter}{row_idx}/({birim_col_letter}{row_idx}*0.15), 0))"
                                )

                        if guncel_setup_col_letter and onerilen_adet_col_letter:
                            for row_idx in range(2, ws.max_row + 1):
                                ws[f"{guncel_setup_col_letter}{row_idx}"] = (
                                    f"={hazirlik_col_letter}{row_idx}/({onerilen_adet_col_letter}{row_idx}*{birim_col_letter}{row_idx})"
                                )
                                ws[f"{guncel_setup_col_letter}{row_idx}"].number_format = "0%"

                        # --- YENİ EKLENTİ: Setup Yükü Veri Çubukları ve %15 Uyarıları ---
                        setup_databar = DataBarRule(
                            start_type="num",
                            start_value=0,
                            end_type="num",
                            end_value=1,
                            color="FFC000",
                        )
                        setup_alert = CellIsRule(
                            operator="greaterThan",
                            formula=["0.15"],
                            font=Font(color="FFFF0000", bold=True),
                        )

                        if setup_yuku_col_letter:
                            ws.conditional_formatting.add(
                                f"{setup_yuku_col_letter}2:{setup_yuku_col_letter}{ws.max_row}",
                                setup_databar,
                            )
                            ws.conditional_formatting.add(
                                f"{setup_yuku_col_letter}2:{setup_yuku_col_letter}{ws.max_row}",
                                setup_alert,
                            )

                        if guncel_setup_col_letter:
                            ws.conditional_formatting.add(
                                f"{guncel_setup_col_letter}2:{guncel_setup_col_letter}{ws.max_row}",
                                setup_databar,
                            )
                            ws.conditional_formatting.add(
                                f"{guncel_setup_col_letter}2:{guncel_setup_col_letter}{ws.max_row}",
                                setup_alert,
                            )

                    # --- YENİ EKLENTİ: Üretim Takip'te üretilenleri istasyonlarda yeşile boyama ---
                    if dk_col_letter:
                        green_fill_row = PatternFill(
                            start_color="92D050", end_color="92D050", fill_type="solid"
                        )
                        green_font_row = Font(color="006100", bold=True)
                        cf_formula = [
                            f"COUNTIF('Üretim Takip'!$I$2:$I${global_max_row}, ${dk_col_letter}2)>0"
                        ]
                        green_rule = FormulaRule(
                            formula=cf_formula,
                            stopIfTrue=False,
                            fill=green_fill_row,
                            font=green_font_row,
                        )
                        ws.conditional_formatting.add(f"A2:Z{max(ws.max_row, 1000)}", green_rule)

                    # --- YENİ EKLENTİ: TOPLAM KAPASİTE / İŞ GÜNÜ HESAPLAMASI ---
                    if toplam_sure_col_letter:
                        # get_station_capacity_settings fonksiyonundan değerleri al
                        gunluk_saat, ozel_makine_sayisi = get_station_capacity_settings(ws_name)
                        if ozel_makine_sayisi is not None:
                            makine_sayisi = ozel_makine_sayisi
                        else:
                            makine_sayisi = machine_counts.get(ws_name, 1)

                        last_row = ws.max_row

                        # İstasyon Sayısı
                        mach_row = last_row + 2
                        ws[f"A{mach_row}"] = "İSTASYON SAYISI"
                        ws[f"A{mach_row}"].font = Font(bold=True)
                        ws[f"{toplam_sure_col_letter}{mach_row}"] = makine_sayisi
                        ws[f"{toplam_sure_col_letter}{mach_row}"].font = Font(bold=True)

                        # Toplam Saat
                        hour_row = mach_row + 1
                        ws[f"A{hour_row}"] = "TOPLAM SÜRE (SAAT)"
                        ws[f"A{hour_row}"].font = Font(bold=True)
                        ws[f"{toplam_sure_col_letter}{hour_row}"] = (
                            f"=SUM({toplam_sure_col_letter}2:{toplam_sure_col_letter}{last_row})/3600"
                        )
                        ws[f"{toplam_sure_col_letter}{hour_row}"].font = Font(bold=True)
                        ws[f"{toplam_sure_col_letter}{hour_row}"].number_format = "#,##0.00"

                        # Toplam İş Günü (Saat / (gunluk_saat * İstasyon Sayısı))
                        day_row = hour_row + 1
                        ws[f"A{day_row}"] = f"TOPLAM İŞ GÜNÜ ({gunluk_saat} Saat/Gün)"
                        ws[f"A{day_row}"].font = Font(bold=True)
                        ws[f"{toplam_sure_col_letter}{day_row}"] = (
                            f"={toplam_sure_col_letter}{hour_row}/({gunluk_saat}*{toplam_sure_col_letter}{mach_row})"
                        )
                        ws[f"{toplam_sure_col_letter}{day_row}"].font = Font(bold=True)
                        ws[f"{toplam_sure_col_letter}{day_row}"].number_format = "#,##0.00"

                        # Dolgu ve kenarlık
                        fill_color = PatternFill(
                            start_color="D9EAD3", end_color="D9EAD3", fill_type="solid"
                        )
                        for r in range(mach_row, day_row + 1):
                            ws[f"A{r}"].fill = fill_color
                            ws[f"{toplam_sure_col_letter}{r}"].fill = fill_color

                elif ws_name == "Üretim Takip":
                    kaynak_col_idx = None
                    kod_col_idx = None
                    uretilecek_col_idx = None
                    uretilen_col_idx = None
                    kalan_col_idx = None
                    tamamlanma_col_idx = None
                    for col in ws.columns:
                        val = col[0].value
                        if not val:
                            continue
                        val_str = str(val).strip()
                        if val_str == "KAYNAK DOSYA":
                            kaynak_col_idx = col[0].column_letter
                        if val_str == "Kod":
                            kod_col_idx = col[0].column_letter
                        if val_str == "Üretilecek Miktar":
                            uretilecek_col_idx = col[0].column_letter
                        if val_str == "Üretilen Miktar":
                            uretilen_col_idx = col[0].column_letter
                        if val_str == "Kalan Miktar":
                            kalan_col_idx = col[0].column_letter
                        if val_str == "Tamamlanma (%)":
                            tamamlanma_col_idx = col[0].column_letter

                    if (
                        kod_col_idx
                        and uretilecek_col_idx
                        and uretilen_col_idx
                        and kalan_col_idx
                        and tamamlanma_col_idx
                    ):
                        wb.ut_kod_col = kod_col_idx
                        wb.ut_uretilen_col = uretilen_col_idx
                        ws["I1"] = "ÜRETİLEN KOD"
                        ws["J1"] = "ÜRETİM ADEDİ"
                        ws["K1"] = "FAZLA ÜRETİM"
                        ws["I1"].font = Font(bold=True)
                        ws["J1"].font = Font(bold=True)
                        ws["K1"].font = Font(bold=True)

                        for row_idx in range(2, ws.max_row + 1):
                            c_kod = f"{kod_col_idx}{row_idx}"
                            c_uretilecek = f"{uretilecek_col_idx}{row_idx}"
                            c_uretilen = f"{uretilen_col_idx}{row_idx}"

                            formula_uretilen = (
                                f"=MAX(0, MIN({c_uretilecek}, "
                                f"IF(COUNTIF($I$2:$I${global_max_row}, {c_kod})>0, IF(SUMIF($I$2:$I${global_max_row}, {c_kod}, $J$2:$J${global_max_row})>0, SUMIF($I$2:$I${global_max_row}, {c_kod}, $J$2:$J${global_max_row}), 9999999), 0) "
                                f"- (SUMIFS(${uretilecek_col_idx}$2:${uretilecek_col_idx}{row_idx}, ${kod_col_idx}$2:${kod_col_idx}{row_idx}, {c_kod}) - {c_uretilecek})"
                                f"))"
                            )
                            ws[c_uretilen] = formula_uretilen
                            ws[f"{kalan_col_idx}{row_idx}"] = f"={c_uretilecek}-{c_uretilen}"
                            ws[f"{tamamlanma_col_idx}{row_idx}"] = (
                                f"=IF({c_uretilecek}>0, {c_uretilen}/{c_uretilecek}, 0)"
                            )
                            ws[f"{tamamlanma_col_idx}{row_idx}"].number_format = "0.00%"

                        # YENİ: Fazla Üretim hesaplamaları (K sütunu)
                        for row_idx in range(2, max(ws.max_row + 500, 1000)):
                            ws[f"K{row_idx}"] = (
                                f'=IF(I{row_idx}<>"", MAX(0, SUMIF($I$2:$I${global_max_row}, I{row_idx}, $J$2:$J${global_max_row}) - SUMIF(${kod_col_idx}$2:${kod_col_idx}${global_max_row}, I{row_idx}, ${uretilecek_col_idx}$2:${uretilecek_col_idx}${global_max_row})), "")'
                            )

                        if kaynak_col_idx:
                            # YENİ: Kaynakları virgülle ayıran ve benzersiz kodları toplayan sistem
                            source_to_codes = {}
                            for row_idx in range(2, ws.max_row + 1):
                                k_val = ws[f"{kaynak_col_idx}{row_idx}"].value
                                c_val = ws[f"{kod_col_idx}{row_idx}"].value
                                if k_val and c_val:
                                    k_list = [str(k).strip() for k in str(k_val).split(",")]
                                    for kaynak in k_list:
                                        if kaynak:
                                            # Clean floating point suffixes like .0
                                            if kaynak.endswith(".0"):
                                                kaynak = kaynak[:-2]
                                            if kaynak not in source_to_codes:
                                                source_to_codes[kaynak] = set()
                                            source_to_codes[kaynak].add(str(c_val).strip())

                            unique_kaynaklar = sorted(list(source_to_codes.keys()))

                            ws["M1"] = "DOSYA BAZLI TAKİP"
                            ws["N1"] = "TOPLAM (Kalem)"
                            ws["O1"] = "HAZIR (Kalem)"
                            ws["P1"] = "EKSİK (Kalem)"
                            ws["Q1"] = "GENEL TAMAMLANMA (%)"

                            for col in ["M", "N", "O", "P", "Q"]:
                                ws[f"{col}1"].font = Font(bold=True)

                            for i, kaynak in enumerate(unique_kaynaklar, start=2):
                                k_str = str(kaynak).strip()
                                if k_str.endswith(".0"):
                                    k_str = k_str[:-2]
                                ws[f"M{i}"] = k_str

                                orig_kalem = 0
                                if file_totals and kaynak in file_totals:
                                    orig_kalem = file_totals[kaynak]

                                # Virgüllü isimlerde formülün patlamaması için wildcard (yıldız) kullanımı
                                sumif_uretilecek = f'SUMIF({kaynak_col_idx}$2:{kaynak_col_idx}${global_max_row}, "*"&M{i}&"*", {uretilecek_col_idx}$2:{uretilecek_col_idx}${global_max_row})'
                                sumif_uretilen = f'SUMIF({kaynak_col_idx}$2:{kaynak_col_idx}${global_max_row}, "*"&M{i}&"*", {uretilen_col_idx}$2:{uretilen_col_idx}${global_max_row})'

                                # Dinamik Eksik ve Hazır kalem hesaplamaları (COUNTIFS ve formül ile)
                                ws[f"P{i}"] = (
                                    f'=COUNTIFS({kaynak_col_idx}$2:{kaynak_col_idx}${global_max_row}, "*"&M{i}&"*", {kalan_col_idx}$2:{kalan_col_idx}${global_max_row}, ">0")'
                                )

                                if orig_kalem > 0:
                                    ws[f"N{i}"] = orig_kalem
                                    ws[f"O{i}"] = f"=MAX(0, N{i}-P{i})"
                                    ws[f"Q{i}"] = (
                                        f"=IF(N{i}>0, (O{i} + P{i} * IF({sumif_uretilecek}>0, {sumif_uretilen}/{sumif_uretilecek}, IF(P{i}>0, 0, 1))) / N{i}, 0)"
                                    )
                                    ws[f"Q{i}"].number_format = "0.00%"
                                else:
                                    ws[f"N{i}"] = (
                                        f'=COUNTIF({kaynak_col_idx}$2:{kaynak_col_idx}${global_max_row}, "*"&M{i}&"*")'
                                    )
                                    ws[f"O{i}"] = f"=MAX(0, N{i}-P{i})"
                                    ws[f"Q{i}"] = (
                                        f"=IF({sumif_uretilecek}>0, {sumif_uretilen}/{sumif_uretilecek}, 0)"
                                    )
                                    ws[f"Q{i}"].number_format = "0.00%"

                        ws.column_dimensions["I"].width = 18
                        ws.column_dimensions["J"].width = 16
                        ws.column_dimensions["K"].width = 18
                        ws.column_dimensions["M"].width = 28
                        ws.column_dimensions["N"].width = 16
                        ws.column_dimensions["O"].width = 16
                        ws.column_dimensions["P"].width = 16
                        ws.column_dimensions["Q"].width = 18
                        ws.column_dimensions["R"].width = 18
                        ws.column_dimensions["S"].width = 25

                        # --- YENİ EKLENTİ: Tamamlanma Yüzdeleri için Veri Çubukları (Data Bars) ---
                        databar_green = DataBarRule(
                            start_type="num",
                            start_value=0,
                            end_type="num",
                            end_value=1,
                            color="5CB85C",
                        )
                        databar_blue = DataBarRule(
                            start_type="num",
                            start_value=0,
                            end_type="num",
                            end_value=1,
                            color="638EC6",
                        )

                        if tamamlanma_col_idx:
                            ws.conditional_formatting.add(
                                f"{tamamlanma_col_idx}2:{tamamlanma_col_idx}{ws.max_row}",
                                databar_green,
                            )

                        if kaynak_col_idx and unique_kaynaklar:
                            q_range = f"Q2:Q{len(unique_kaynaklar) + 1}"
                            ws.conditional_formatting.add(q_range, databar_blue)

                            # YENİ: Renkli Koşullu Biçimlendirme (Kırmızı/Sarı/Yeşil)
                            red_fill_cell = PatternFill(
                                start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                            )
                            yellow_fill_cell = PatternFill(
                                start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
                            )
                            green_fill_cell = PatternFill(
                                start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
                            )

                            ws.conditional_formatting.add(
                                q_range,
                                CellIsRule(
                                    operator="lessThan", formula=["0.2"], fill=red_fill_cell
                                ),
                            )
                            ws.conditional_formatting.add(
                                q_range,
                                CellIsRule(
                                    operator="between",
                                    formula=["0.2", "0.8"],
                                    fill=yellow_fill_cell,
                                ),
                            )
                            ws.conditional_formatting.add(
                                q_range,
                                CellIsRule(
                                    operator="greaterThan", formula=["0.8"], fill=green_fill_cell
                                ),
                            )

                        # --- YENİ EKLENTİ: HATALI KOD GİRİŞİNDE KIRMIZI UYARI (Koşullu Biçimlendirme) ---
                        if kod_col_idx:
                            red_fill = PatternFill(
                                start_color="FFFF0000", end_color="FFFF0000", fill_type="solid"
                            )
                            cf_rule = FormulaRule(
                                formula=[
                                    f'AND($I2<>"", COUNTIF(${kod_col_idx}$2:${kod_col_idx}${global_max_row}, $I2)=0)'
                                ],
                                stopIfTrue=True,
                                fill=red_fill,
                            )
                            ws.conditional_formatting.add(
                                f"I2:I{max(ws.max_row + 500, 1000)}", cf_rule
                            )

                        # Üretim Takip sayfasında A-H sütunlarını kesin olarak gizle
                        for col_letter in ["A", "B", "C", "D", "E", "F", "G", "H"]:
                            ws.column_dimensions[col_letter].hidden = True

                    # YENİ EKLENTİ: En alt satıra TOPLAM SÜRE ve TOPLAM SAAT ekleme
                    if toplam_sure_col_letter:
                        last_data_row = ws.max_row

                        # Toplam Saniye Satırı
                        sum_saniye_row = last_data_row + 2
                        ws[f"A{sum_saniye_row}"] = "GENEL TOPLAM SÜRE (SANİYE)"
                        ws[f"A{sum_saniye_row}"].font = Font(bold=True)
                        ws[f"{toplam_sure_col_letter}{sum_saniye_row}"] = (
                            f"=SUM({toplam_sure_col_letter}2:{toplam_sure_col_letter}{last_data_row})"
                        )
                        ws[f"{toplam_sure_col_letter}{sum_saniye_row}"].font = Font(bold=True)
                        ws[f"{toplam_sure_col_letter}{sum_saniye_row}"].number_format = "#,##0"

                        # Toplam Saat Satırı
                        sum_saat_row = last_data_row + 3
                        ws[f"A{sum_saat_row}"] = "TOPLAM İŞ SAATİ (SAAT)"
                        ws[f"A{sum_saat_row}"].font = Font(bold=True)
                        ws[f"{toplam_sure_col_letter}{sum_saat_row}"] = (
                            f"={toplam_sure_col_letter}{sum_saniye_row}/3600"
                        )
                        ws[f"{toplam_sure_col_letter}{sum_saat_row}"].font = Font(bold=True)
                        ws[f"{toplam_sure_col_letter}{sum_saat_row}"].number_format = "#,##0.00"

                        # Toplam İş Günü Satırı
                        sum_gun_row = last_data_row + 4
                        ws[f"A{sum_gun_row}"] = "TOPLAM İŞ GÜNÜ (GÜN)"
                        ws[f"A{sum_gun_row}"].font = Font(bold=True)

                        # İstasyonlara göre günlük çalışma saati kapasitesi
                        gunluk_saat, ozel_makine_sayisi = get_station_capacity_settings(ws_name)
                        makine_sayisi = ozel_makine_sayisi if ozel_makine_sayisi is not None else 1
                        gunluk_kapasite = gunluk_saat * makine_sayisi

                        ws[f"{toplam_sure_col_letter}{sum_gun_row}"] = (
                            f"={toplam_sure_col_letter}{sum_saat_row}/{gunluk_kapasite}"
                        )
                        ws[f"{toplam_sure_col_letter}{sum_gun_row}"].font = Font(bold=True)
                        ws[f"{toplam_sure_col_letter}{sum_gun_row}"].number_format = "#,##0.00"

            # Tüm hücrelere kenarlık ve başlık kalınlığı uygula
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
            no_border = Border()
            bold_font = Font(bold=True)
            right_alignment = Alignment(horizontal="right")

            # Hariç tutulan kodları JSON'dan yükle
            excluded_codes = set()
            db_dir = os.path.join(os.getcwd(), settings.db_dir_name)
            json_path = os.path.join(db_dir, "haric_tutulacak_parcalar.json")
            if os.path.exists(json_path):
                try:
                    with open(json_path, encoding="utf-8") as f:
                        data = json.load(f)
                        if isinstance(data, list):
                            excluded_codes = {str(k).strip().upper() for k in data}
                except Exception:
                    pass

            # Performans optimizasyonu: matched_tim_relations listesini aramaları hızlandırmak için önceden grupla/indeksle
            tim_relations_by_parent = {}
            tim_relations_by_child = {}

            for rel in matched_tim_relations:
                r_kaynak = str(rel["row_ref"].get("KAYNAK DOSYA", kaynak_metin_kisa)).strip()
                r_parent = str(rel["p_row"].get("Kod", rel["p_row"].get(sira_col, ""))).strip()
                r_child = str(rel["parca_kodu"]).strip()

                # Parent bazlı gruplama
                key_p = (r_kaynak, r_parent)
                if key_p not in tim_relations_by_parent:
                    tim_relations_by_parent[key_p] = []
                tim_relations_by_parent[key_p].append(rel)

                # Child bazlı birebir eşleme
                key_c = (r_kaynak, r_parent, r_child)
                tim_relations_by_child[key_c] = rel

            purple_fill = PatternFill(start_color="F2EBF9", end_color="F2EBF9", fill_type="solid")
            purple_font = Font(color="6C3483")

            for sheet_name in wb.sheetnames:
                ws_format = wb[sheet_name]

                # 1. Başlık satırını kalın yap
                for cell in ws_format[1]:
                    cell.font = bold_font

                # Sütunlarda "Kod" veya "Hammadde Kod" araması yap
                code_col_letter = None
                code_col_letters = set()
                for col in ws_format.columns:
                    val = str(col[0].value).strip().lower()
                    if (
                        val
                        in [
                            "kod",
                            "alt parça kodu",
                            "alt parca kodu",
                            "hammadde kod",
                            "hammadde kodu",
                        ]
                        and code_col_letter is None
                    ):
                        code_col_letter = col[0].column_letter

                    n_col = norm_col(val)
                    if any(k in n_col for k in ["kod", "kodu"]):
                        code_col_letters.add(col[0].column_letter)

                # Mor renge boyanacak satırların listesini bul
                purple_rows = set()
                if code_col_letter:
                    for row_idx in range(2, ws_format.max_row + 1):
                        cell_val = (
                            str(ws_format[f"{code_col_letter}{row_idx}"].value).strip().upper()
                        )
                        if cell_val in excluded_codes:
                            purple_rows.add(row_idx)

                # 2. Hücreleri biçimlendir ve kenarlık ekle
                if "ZLEME" in sheet_name.upper():
                    current_kaynak_dosya = None
                    current_group_code = None
                    use_alt_fill = False

                    alt_fill = PatternFill(
                        start_color="D2E3FC", end_color="D2E3FC", fill_type="solid"
                    )
                    white_fill = PatternFill(
                        start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
                    )

                    source_file_border = Border(
                        left=Side(style="thin"),
                        right=Side(style="thin"),
                        top=Side(style="medium", color="000000"),
                        bottom=Side(style="thin"),
                    )

                    group_top_border = Border(
                        left=Side(style="thin"),
                        right=Side(style="thin"),
                        top=Side(style="thin", color="888888"),
                        bottom=Side(style="thin"),
                    )

                    for row_idx in range(2, ws_format.max_row + 1):
                        kaynak_dosya = str(ws_format[f"A{row_idx}"].value).strip()
                        parent_code = str(ws_format[f"B{row_idx}"].value).strip()

                        is_source_start = False
                        is_group_start = False

                        if kaynak_dosya != current_kaynak_dosya:
                            current_kaynak_dosya = kaynak_dosya
                            current_group_code = parent_code
                            use_alt_fill = not use_alt_fill
                            is_source_start = True
                        elif parent_code != current_group_code:
                            current_group_code = parent_code
                            is_group_start = True

                        for col_letter in ["A", "B", "C", "D", "E", "F", "G", "H"]:
                            cell = ws_format[f"{col_letter}{row_idx}"]

                            if row_idx in purple_rows:
                                cell.fill = purple_fill
                                cell.font = purple_font
                            else:
                                if use_alt_fill:
                                    cell.fill = alt_fill
                                else:
                                    cell.fill = white_fill

                            if is_source_start and row_idx > 2:
                                cell.border = source_file_border
                            elif is_group_start and row_idx > 2:
                                cell.border = group_top_border
                            else:
                                cell.border = thin_border

                            if col_letter in ["B", "D"]:
                                cell.alignment = Alignment(horizontal="right")

                    # --- YENİ EKLENTİ: BENZERSİZ ÜST MONTAJ CANLI TAKİP YAN TABLOSU ---
                    ws_format["K1"] = "Kaynak Dosya"
                    ws_format["L1"] = "Üst Montaj Kodu"
                    ws_format["M1"] = "Üst Montaj Adı"
                    ws_format["N1"] = "Gereken Çeşit"
                    ws_format["O1"] = "Tamamlanan Çeşit"
                    ws_format["P1"] = "Tamamlanma Oranı (%)"
                    ws_format["Q1"] = "Toplam Parça Adedi"
                    ws_format["R1"] = "Ek Toplanabilir (Set)"

                    for col_letter in ["K", "L", "M", "N", "O", "P", "Q", "R"]:
                        ws_format[f"{col_letter}1"].font = bold_font

                    # Benzersiz Üst Montajları sırayla topla
                    unique_parents = []
                    seen_parents = set()
                    for row_idx in range(2, ws_format.max_row + 1):
                        k_dosya = ws_format[f"A{row_idx}"].value
                        p_kod = ws_format[f"B{row_idx}"].value
                        p_ad = ws_format[f"C{row_idx}"].value
                        if k_dosya is not None and p_kod is not None:
                            key = (str(k_dosya).strip(), str(p_kod).strip())
                            if key not in seen_parents:
                                seen_parents.add(key)
                                unique_parents.append({"kaynak": k_dosya, "kod": p_kod, "ad": p_ad})

                    current_kaynak_dosya_side = None
                    use_alt_fill_side = False

                    parent_stats = getattr(wb, "parent_stats", {})
                    for i, parent in enumerate(unique_parents):
                        write_row = i + 2
                        ws_format[f"K{write_row}"] = parent["kaynak"]
                        ws_format[f"L{write_row}"] = parent["kod"]
                        ws_format[f"M{write_row}"] = parent["ad"]

                        # COUNTIFS ve oransal formüller (Çeşit bazlı, stok dahil)
                        kaynak_str = str(parent["kaynak"]).strip()
                        kod_str = str(parent["kod"]).strip()
                        key = (kaynak_str, kod_str)
                        stats = parent_stats.get(
                            key, {"total": set(), "missing": set(), "total_qty": 0.0}
                        )
                        total_count = len(stats["total"])
                        missing_count = len(stats["missing"])
                        in_stock_count = total_count - missing_count
                        total_qty = stats.get("total_qty", 0.0)

                        if total_count == 0:
                            ws_format[f"N{write_row}"] = (
                                f"=COUNTIFS(A:A, K{write_row}, B:B, L{write_row})"
                            )
                            ws_format[f"O{write_row}"] = (
                                f"=COUNTIFS(A:A, K{write_row}, B:B, L{write_row}, H:H, 1)"
                            )
                            ws_format[f"Q{write_row}"] = (
                                f"=SUMIFS(F:F, A:A, K{write_row}, B:B, L{write_row})"
                            )
                        else:
                            ws_format[f"N{write_row}"] = total_count
                            if in_stock_count > 0:
                                ws_format[f"O{write_row}"] = (
                                    f"={in_stock_count}+COUNTIFS(A:A, K{write_row}, B:B, L{write_row}, H:H, 1)"
                                )
                            else:
                                ws_format[f"O{write_row}"] = (
                                    f"=COUNTIFS(A:A, K{write_row}, B:B, L{write_row}, H:H, 1)"
                                )

                            if isinstance(total_qty, float) and total_qty.is_integer():
                                total_qty = int(total_qty)
                            ws_format[f"Q{write_row}"] = total_qty

                        ws_format[f"P{write_row}"] = (
                            f"=IF(N{write_row}>0, O{write_row}/N{write_row}, 0)"
                        )
                        ws_format[f"P{write_row}"].number_format = "0%"

                        # Canlı Ek Toplanabilir (Set) Formülü
                        in_stock_limits = []
                        parent_relations = tim_relations_by_parent.get(key, [])
                        for rel in parent_relations:
                            if rel.get("allocated_qty", 0.0) <= 0:
                                c_kod = str(rel["parca_kodu"]).strip().upper()
                                val_eldeki = stok_dict_kullanilabilir.get(c_kod, 0.0)
                                val_ihtiyac = total_required_qty.get(c_kod, 0.0)

                                p_qty = 1.0
                                for qc in ["Miktar", "Adet", "Qty", "Quantity"]:
                                    if qc in rel["p_row"]:
                                        try:
                                            p_qty = float(
                                                str(rel["p_row"].get(qc, 1)).replace(",", ".")
                                            )
                                            break
                                        except ValueError:
                                            pass
                                c_design_qty = rel.get("design_qty", rel["orig_qty"])
                                birim_ihtiyac = c_design_qty / p_qty if p_qty > 0 else 0.0
                                if birim_ihtiyac > 0:
                                    limit = int(max(0.0, val_eldeki - val_ihtiyac) / birim_ihtiyac)
                                    in_stock_limits.append(limit)

                        in_stock_tim_limit = min(in_stock_limits) if in_stock_limits else None

                        if in_stock_tim_limit is not None:
                            ws_format[f"R{write_row}"] = (
                                f"=IF(P{write_row}=1, "
                                f"IF(COUNTIFS(A:A, K{write_row}, B:B, L{write_row})>0, "
                                f"MIN({in_stock_tim_limit}, _xlfn.MINIFS(I:I, A:A, K{write_row}, B:B, L{write_row})), "
                                f"{in_stock_tim_limit}), 0)"
                            )
                        else:
                            ws_format[f"R{write_row}"] = (
                                f"=IF(P{write_row}=1, _xlfn.MINIFS(I:I, A:A, K{write_row}, B:B, L{write_row}), 0)"
                            )

                        is_source_start_side = False
                        if parent["kaynak"] != current_kaynak_dosya_side:
                            current_kaynak_dosya_side = parent["kaynak"]
                            use_alt_fill_side = not use_alt_fill_side
                            is_source_start_side = True

                        for col_letter in ["K", "L", "M", "N", "O", "P", "Q", "R"]:
                            cell = ws_format[f"{col_letter}{write_row}"]
                            if use_alt_fill_side:
                                cell.fill = alt_fill
                            else:
                                cell.fill = white_fill

                            if is_source_start_side and write_row > 2:
                                cell.border = source_file_border
                            else:
                                cell.border = thin_border

                            if col_letter in ["L", "R"]:
                                cell.alignment = Alignment(horizontal="right")

                    # Sütun genişliklerini sabitle
                    ws_format.column_dimensions["K"].width = 15
                    ws_format.column_dimensions["L"].width = 18
                    ws_format.column_dimensions["M"].width = 30
                    ws_format.column_dimensions["N"].width = 16
                    ws_format.column_dimensions["O"].width = 16
                    ws_format.column_dimensions["P"].width = 20
                    ws_format.column_dimensions["Q"].width = 20
                    ws_format.column_dimensions["R"].width = 22
                else:
                    for row in ws_format.iter_rows():
                        if row[0].row == 1:
                            continue
                        row_idx = row[0].row
                        is_purple = row_idx in purple_rows

                        for cell in row:
                            if sheet_name == "Üretim Takip" and cell.column_letter in ["H", "L"]:
                                cell.border = no_border
                            else:
                                cell.border = thin_border

                            if is_purple:
                                cell.fill = purple_fill
                                cell.font = purple_font

                            if cell.column_letter in code_col_letters:
                                cell.alignment = Alignment(horizontal="right")

            for sheet_name in wb.sheetnames:
                if "ZLEME" in sheet_name.upper():
                    ws_izleme = wb[sheet_name]
                    alt_kod_col = None
                    gereken_col = None
                    uretilen_col = None
                    oran_col = None
                    for col in list(ws_izleme.columns)[:8]:
                        val = str(col[0].value).strip()
                        if val == "Alt Parça Kodu":
                            alt_kod_col = col[0].column_letter
                        if val == "Gereken Miktar":
                            gereken_col = col[0].column_letter
                        if val == "Üretilen Miktar":
                            uretilen_col = col[0].column_letter
                        if val == "Tamamlanma Oranı (%)":
                            oran_col = col[0].column_letter

                    if (
                        alt_kod_col
                        and uretilen_col
                        and hasattr(wb, "ut_kod_col")
                        and hasattr(wb, "ut_uretilen_col")
                    ):
                        ws_izleme["I1"] = "Alt Parça Limit"
                        ws_izleme["I1"].font = Font(bold=True)
                        main_max_row = 1
                        for r in range(2, ws_izleme.max_row + 1):
                            if ws_izleme.cell(row=r, column=1).value is not None:
                                main_max_row = r

                        for row_idx in range(2, main_max_row + 1):
                            alt_kodu = f"{alt_kod_col}{row_idx}"
                            req_qty = f"{gereken_col}{row_idx}"
                            ws_izleme[f"{uretilen_col}{row_idx}"] = (
                                f"=MAX(0, MIN({req_qty}, "
                                f"SUMIF('Üretim Takip'!{wb.ut_kod_col}:{wb.ut_kod_col}, {alt_kodu}, 'Üretim Takip'!{wb.ut_uretilen_col}:{wb.ut_uretilen_col}) - "
                                f"SUMIFS(${uretilen_col}$1:${uretilen_col}{row_idx-1}, ${alt_kod_col}$1:${alt_kod_col}{row_idx-1}, {alt_kodu})"
                                f"))"
                            )
                            ws_izleme[f"{oran_col}{row_idx}"] = (
                                f"=IF({req_qty}>0, {uretilen_col}{row_idx}/{req_qty}, 0)"
                            )
                            ws_izleme[f"{oran_col}{row_idx}"].number_format = "0%"

                            # Canlı Alt Parça Ek Set Limiti Formülü
                            kaynak_str = str(ws_izleme.cell(row=row_idx, column=1).value).strip()
                            parent_str = str(ws_izleme.cell(row=row_idx, column=2).value).strip()
                            child_str = str(ws_izleme.cell(row=row_idx, column=4).value).strip()

                            key_child = (kaynak_str, parent_str, child_str)
                            matching_rel = tim_relations_by_child.get(key_child)

                            birim_ihtiyac = 0.0
                            if matching_rel:
                                p_qty = 1.0
                                for qc in ["Miktar", "Adet", "Qty", "Quantity", miktar_col]:
                                    if qc in matching_rel["p_row"]:
                                        try:
                                            p_qty = float(
                                                str(matching_rel["p_row"].get(qc, 1)).replace(
                                                    ",", "."
                                                )
                                            )
                                            break
                                        except ValueError:
                                            pass
                                c_design_qty = matching_rel.get(
                                    "design_qty", matching_rel["orig_qty"]
                                )
                                birim_ihtiyac = c_design_qty / p_qty if p_qty > 0 else 0.0

                            if birim_ihtiyac > 0:
                                val_eldeki = stok_dict_kullanilabilir.get(child_str, 0.0)
                                val_ihtiyac = total_required_qty.get(child_str, 0.0)
                                sumif_str = (
                                    f"SUMIF('Üretim Takip'!I:I, {alt_kodu}, 'Üretim Takip'!J:J)"
                                )
                                ws_izleme[f"I{row_idx}"] = (
                                    f"=IF({uretilen_col}{row_idx}>={req_qty}, "
                                    f"INT(MAX(0, {val_eldeki} + {sumif_str} - {val_ihtiyac}) / {birim_ihtiyac}), 0)"
                                )
                            else:
                                ws_izleme[f"I{row_idx}"] = 0

                        ws_izleme.column_dimensions["I"].hidden = True

                        # Tamamlanmış montaj gruplarını (Gereken Toplamı == Üretilen Toplamı) otomatik yeşile boya
                        green_fill = PatternFill(
                            start_color="92D050", end_color="92D050", fill_type="solid"
                        )
                        green_font = Font(color="006100", bold=True)
                        cf_formula = f'AND($A2<>"", SUMIFS(${gereken_col}:${gereken_col}, $A:$A, $A2, $B:$B, $B2)=SUMIFS(${uretilen_col}:${uretilen_col}, $A:$A, $A2, $B:$B, $B2))'
                        green_rule = FormulaRule(
                            formula=[cf_formula], stopIfTrue=False, fill=green_fill, font=green_font
                        )
                        ws_izleme.conditional_formatting.add(f"A2:H{main_max_row}", green_rule)

                        # Tamamlanma Oranı (%) sütununa yeşil veri çubuğu (Data Bar) ekle
                        databar_izleme = DataBarRule(
                            start_type="num",
                            start_value=0,
                            end_type="num",
                            end_value=1,
                            color="5CB85C",
                        )
                        ws_izleme.conditional_formatting.add(
                            f"{oran_col}2:{oran_col}{main_max_row}", databar_izleme
                        )

                        # Yan tablonun Tamamlanma Oranı (%) sütununa (P sütunu) mavi veri çubuğu ekle
                        side_max_row = 1
                        for r in range(2, ws_izleme.max_row + 1):
                            if ws_izleme.cell(row=r, column=11).value is not None:
                                side_max_row = r

                        databar_izleme_side = DataBarRule(
                            start_type="num",
                            start_value=0,
                            end_type="num",
                            end_value=1,
                            color="5CB85C",
                        )
                        ws_izleme.conditional_formatting.add(
                            f"P2:P{side_max_row}", databar_izleme_side
                        )

                        # Yan tablonun tamamlanan gruplarını (%100 tamamlanma oranı) yeşile boya
                        green_fill_side = PatternFill(
                            start_color="92D050", end_color="92D050", fill_type="solid"
                        )
                        green_font_side = Font(color="006100", bold=True)
                        cf_formula_side = 'AND($K2<>"", $P2=1)'
                        green_rule_side = FormulaRule(
                            formula=[cf_formula_side],
                            stopIfTrue=False,
                            fill=green_fill_side,
                            font=green_font_side,
                        )
                        ws_izleme.conditional_formatting.add(f"K2:Q{side_max_row}", green_rule_side)

            # Tüm oluşturulan excel sayfalarının sütun genişliklerini içeriğe/yazılara göre dinamik ayarla
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                is_izleme = "ZLEME" in sheet_name.upper()
                for col in ws.columns:
                    col_letter = col[0].column_letter
                    if ws.column_dimensions[col_letter].hidden:
                        continue

                    max_len = 0
                    for cell in col:
                        val = cell.value
                        if val is not None:
                            val_str = str(val)
                            # Eğer formül ise, formülün sonucunun tahmini uzunluğunu kullanalım
                            if val_str.startswith("="):
                                if any(
                                    x in val_str.upper()
                                    for x in ["COUNTIFS", "SUMIF", "SUMIFS", "IF", "/"]
                                ):
                                    val_len = 6
                                else:
                                    val_len = 10
                            else:
                                val_len = len(val_str)
                            if val_len > max_len:
                                max_len = val_len

                    # İzleme sayfasının yan tablosu için biraz daha geniş sınırlar tanıyabiliriz
                    padding = (
                        5 if is_izleme and col_letter in ["K", "L", "M", "N", "O", "P", "Q"] else 4
                    )
                    adjusted_width = max(max_len + padding, 12)
                    ws.column_dimensions[col_letter].width = adjusted_width

            # Final Kaydetme
            console.print(
                "  [cyan]⏳ Excel dosyası kaydediliyor (Bu işlem biraz zaman alabilir)...[/cyan]"
            )
            wb.save(out_path)
            wb.close()
            if gizlenen_adet > 0:
                console.print(
                    f"  [cyan]👁️  İstediğiniz {gizlenen_adet} özel sütun Excel'deki TÜM SAYFALARDA gizlendi.[/cyan]"
                )
        except Exception as e:
            console.print(
                f"  [yellow]Uyarı:[/yellow] Sütun biçimlendirme sırasında hata oluştu: {e}"
            )

        # --- YENİ EKLENTİ: OTOMATİK REÇETE AĞACI OLUŞTURMA ---
        if input_path:
            try:
                console.print(
                    "\n[bold cyan]=== OTOMATİK REÇETE AĞACI OLUŞTURULUYOR ===[/bold cyan]"
                )
                from recipe_automation.services.bom_tree import build_bom_tree
                from recipe_automation.utils.excel_io import find_excel_files

                db_dir = os.path.join(os.getcwd(), settings.db_dir_name)
                prio_mapping = load_priority_mapping(db_dir)

                def get_file_priority_and_name(fpath, mapping):
                    base = os.path.splitext(os.path.basename(fpath))[0]
                    clean_name = base
                    for prefix in ["Filtered_TIM_", "Filtered_", "TIM_"]:
                        if clean_name.startswith(prefix):
                            clean_name = clean_name[len(prefix) :]
                            break
                    prio = 999999
                    if clean_name in mapping:
                        prio = mapping[clean_name]
                    else:
                        for k, v in mapping.items():
                            if k in clean_name or clean_name in k:
                                if v < prio:
                                    prio = v
                    return prio, clean_name, base

                def sort_key(fpath):
                    prio, _, base = get_file_priority_and_name(fpath, prio_mapping)
                    return (prio, base.lower())

                excel_files = find_excel_files(input_path)
                if excel_files:
                    # Filtreleme çıktısı olan dosyaları (Filtered_ ile başlayanlar veya grup adı ile başlayanlar) atla
                    group_names = []
                    operasyon_file = os.path.join(
                        os.getcwd(), settings.db_dir_name, "operasyon_gruplari.json"
                    )
                    if os.path.exists(operasyon_file):
                        try:
                            with open(operasyon_file, encoding="utf-8") as f:
                                groups_data = json.load(f)
                                group_names = [str(k).upper().strip() for k in groups_data.keys()]
                        except Exception:
                            pass
                    if not group_names:
                        group_names = ["TIM", "BUKUM", "KAYNAK", "LAZER KESIM", "3D YAZICI"]

                    filtered_excel_files = []
                    for fpath in excel_files:
                        fname = os.path.basename(fpath)
                        if fname.lower().startswith("filtered_"):
                            continue
                        is_output = False
                        for g in group_names:
                            if fname.upper().startswith(g + "_"):
                                is_output = True
                                break
                        if is_output:
                            continue
                        filtered_excel_files.append(fpath)

                    excel_files = sorted(filtered_excel_files, key=sort_key)
                    all_trees = []
                    ut_df = sheet_dfs.get("Üretim Takip")
                    for fpath in excel_files:
                        fname = os.path.basename(fpath)
                        console.print(f"  [cyan]⏳ Reçete ağacı çözümleniyor: {fname}...[/cyan]")
                        df_tree = None
                        for h in [0, 1, 2, 3]:
                            try:
                                candidate = pd.read_excel(fpath, header=h, nrows=5)
                                cols_norm = [
                                    str(c)
                                    .strip()
                                    .lower()
                                    .replace("ı", "i")
                                    .replace("ş", "s")
                                    .replace("ç", "c")
                                    .replace("ğ", "g")
                                    for c in candidate.columns
                                ]
                                if any(c in ["sira", "sira no"] for c in cols_norm):
                                    df_tree = pd.read_excel(fpath, header=h)
                                    break
                            except Exception:
                                continue

                        if df_tree is not None and not df_tree.empty:
                            try:
                                # Üretim Takip sayfasından başlangıç ihtiyaç adetlerini çek
                                initial_reqs = {}
                                if ut_df is not None and not ut_df.empty:
                                    fname_no_ext = os.path.splitext(fname)[0]
                                    if "KAYNAK DOSYA" in ut_df.columns:
                                        file_mask = (
                                            ut_df["KAYNAK DOSYA"]
                                            .astype(str)
                                            .str.strip()
                                            .apply(
                                                lambda x: x == fname_no_ext
                                                or x == fname
                                                or x.startswith(fname_no_ext)
                                            )
                                        )
                                        file_ut_df = ut_df[file_mask]
                                    else:
                                        file_ut_df = ut_df

                                    if (
                                        not file_ut_df.empty
                                        and "Kod" in file_ut_df.columns
                                        and "Üretilecek Miktar" in file_ut_df.columns
                                    ):
                                        for _, row in file_ut_df.iterrows():
                                            k = str(row["Kod"]).strip().upper()
                                            try:
                                                val = float(
                                                    str(row["Üretilecek Miktar"]).replace(",", ".")
                                                )
                                            except Exception:
                                                val = 0.0
                                            if k and k != "NAN" and k != "":
                                                initial_reqs[k] = initial_reqs.get(k, 0.0) + val

                                prio, _, base_name = get_file_priority_and_name(fpath, prio_mapping)
                                displayName = base_name
                                if prio < 999999:
                                    displayName = f"[{prio}] {base_name}"

                                tree_dict = build_bom_tree(
                                    df_tree, filename=displayName, initial_reqs=initial_reqs
                                )

                                # Makine adını çekip displayName'e ekle
                                machine_name = ""
                                if tree_dict:
                                    ad_val = str(tree_dict.get("ad", "")).strip()
                                    if ad_val and ad_val.upper() not in [
                                        "ROOT",
                                        "ANA MAKİNE",
                                        displayName.upper(),
                                        "ANA MAKINE",
                                    ]:
                                        machine_name = ad_val
                                    else:
                                        kod_val = str(tree_dict.get("kod", "")).strip()
                                        if kod_val and kod_val.upper() not in ["ROOT"]:
                                            machine_name = kod_val
                                if machine_name:
                                    displayName = f"{displayName} ({machine_name})"

                                all_trees.append({"name": displayName, "tree": tree_dict})
                            except Exception as e:
                                console.print(
                                    f"  [yellow]Uyarı:[/yellow] {fname} için ağaç oluşturulamadı: {e}"
                                )

                    if all_trees:
                        all_trees_json = json.dumps(all_trees, ensure_ascii=False)
                        template_path = os.path.join(
                            os.path.dirname(__file__), "web", "template.html"
                        )
                        if os.path.exists(template_path):
                            with open(template_path, encoding="utf-8") as f:
                                html_content = f.read()

                            html_content = html_content.replace(
                                "__TREE_DATA_PLACEHOLDER__",
                                json.dumps(all_trees[0]["tree"], ensure_ascii=False),
                            )
                            html_content = html_content.replace(
                                "__ALL_TREES_PLACEHOLDER__", all_trees_json
                            )
                            html_content = html_content.replace(
                                "__FILE_NAME_PLACEHOLDER__", all_trees[0]["name"]
                            )

                            output_dir = os.path.dirname(path)
                            base_name_clean = os.path.splitext(os.path.basename(path))[0]
                            if base_name_clean.startswith("temp_scaled_input_"):
                                base_name_clean = base_name_clean[len("temp_scaled_input_") :]
                            output_name = f"Makine_Agaci_{base_name_clean}.html"
                            if len(excel_files) > 1:
                                output_name = "Makine_Agaci_Receteler.html"

                            out_tree_path = os.path.join(output_dir, output_name)
                            with open(out_tree_path, "w", encoding="utf-8") as f:
                                f.write(html_content)
                            console.print(
                                f"[green]✅ Reçete Ağacı oluşturuldu:[/green] {out_tree_path}"
                            )
            except Exception as e:
                console.print(
                    f"[yellow]Uyarı: Otomatik ağaç oluşturma sırasında hata oluştu:[/yellow] {e}"
                )

        if unassigned_stations:
            console.print(
                "\n[yellow]💡 Bilgi: Aşağıdaki istasyonlar herhangi bir gruba ayrılmadı (Yalnızca 'Tüm Veriler' ana sayfasında varlar):[/yellow]"
            )
            for ust in sorted(unassigned_stations):
                console.print(f"  - [white]{ust}[/white]")
        # ------------------------------------------------

        table = Table(show_header=False, box=None)
        table.add_column("Key", style="cyan")
        table.add_column("Value", style="magenta")
        table.add_row("Filtrelenmiş Exceldeki Kod Sayısı", str(len(filtered_df)))
        table.add_row("Depodaki Toplam Satır", str(len(depo_df)))
        table.add_row("Depoda Bulunan/Kalan Satır", str(len(matched_df)))
        table.add_row("Depodan Silinen Satır Sayısı", str(len(depo_df) - len(matched_df)))

        panel = Panel(
            table,
            title="[bold green]Eşleştirme Raporu[/bold green]",
            border_style="green",
            expand=False,
        )
        console.print(panel)
        console.print(f"[green]✅ Çıktı oluşturuldu:[/green] {out_path}\n")

    except ValueError as e:
        console.print(f"[red]Hata:[/red] {e}")
        raise typer.Exit(1)


@app.command()
def reserve(
    path: str = typer.Argument(..., help="İşlenecek Excel klasörünün yolu (sürükle-bırak desteklenir)"),
) -> None:
    """
    Öncelik sırasına göre klasördeki Excel reçetelerine stok rezervasyonu uygular.
    StokListesi.xlsx verilerini FIFO mantığıyla dağıtır (Seçenek A: kısmi rezervasyon).
    Çıktı dosyaları orijinal adın sonuna '_' eklenerek aynı klasöre kaydedilir.
    """
    from recipe_automation.utils.stock_reserver import load_stok_dict, reserve_single_file
    from recipe_automation.services.sorter import load_priority_mapping

    # ── Yol doğrulama ────────────────────────────────────────────────────────
    if not os.path.isdir(path):
        console.print(f"[red]Hata:[/red] Geçerli bir klasör yolu bekleniyor: {path}")
        raise typer.Exit(1)

    # ── Veritabanı ────────────────────────────────────────────────────────────
    db_dir = os.path.join(os.getcwd(), settings.db_dir_name)
    stok_path = os.path.join(db_dir, "StokListesi.xlsx")
    priority_file = os.path.join(db_dir, "oncelik_sirasi.json")

    # ── Öncelik eşlemesi ─────────────────────────────────────────────────────
    prio_mapping: dict[str, int] = {}
    if os.path.exists(priority_file):
        with open(priority_file, encoding="utf-8") as f:
            prio_mapping = json.load(f)
    else:
        console.print(f"[yellow]Uyarı:[/yellow] {priority_file} bulunamadı. Dosyalar alfabetik sırayla işlenecek.")

    # ── Excel dosyalarını bul ve sırala ──────────────────────────────────────
    all_xlsx = [
        os.path.join(path, f)
        for f in os.listdir(path)
        if f.lower().endswith(".xlsx")
        and not f.startswith("~$")
        and not f.lower().endswith("_.xlsx")  # Zaten işlenmiş _ uzantılıları atla
        and os.path.basename(f) not in [
            "StokListesi.xlsx", "TumRotaBilgileri.xlsx", "ReceteTumRotaListe.xlsx"
        ]
    ]

    if not all_xlsx:
        console.print(f"[red]Hata:[/red] Klasörde işlenecek Excel dosyası bulunamadı: {path}")
        raise typer.Exit(1)

    def get_priority(fpath: str) -> int:
        base = os.path.splitext(os.path.basename(fpath))[0]
        # Olası ön ekleri temizle (Filtered_TIM_ vb.)
        for prefix in ["Filtered_TIM_", "Filtered_", "TIM_"]:
            if base.startswith(prefix):
                base = base[len(prefix):]
                break
        if base in prio_mapping:
            return prio_mapping[base]
        # Kısmi eşleşme dene
        for k, v in prio_mapping.items():
            if k in base or base in k:
                return v
        return 999999  # Önceliği bilinmeyenler sona

    all_xlsx.sort(key=lambda f: (get_priority(f), os.path.basename(f).lower()))

    # ── StokListesi'ni yükle ─────────────────────────────────────────────────
    console.print("\n[bold cyan]=== STOK REZERVASYON SİSTEMİ ===[/bold cyan]")
    if os.path.exists(stok_path):
        console.print(f"[cyan]📦 Stok listesi yükleniyor:[/cyan] {stok_path}")
        stok_dict = load_stok_dict(stok_path)
        console.print(f"[green]✅ {len(stok_dict)} parça kodu için stok bilgisi yüklendi.[/green]\n")
    else:
        console.print(f"[yellow]⚠️  StokListesi.xlsx bulunamadı ({stok_path}).[/yellow]")
        console.print("[yellow]   Stok kontrolü yapılmadan tüm ihtiyaçlar 'Rezerve Edilecek' olarak işaretlenecek.[/yellow]\n")
        stok_dict = {}

    console.print(f"[bold white]{len(all_xlsx)} dosya öncelik sırasına göre işlenecek:[/bold white]")
    for i, f in enumerate(all_xlsx, 1):
        base = os.path.basename(f)
        prio = get_priority(f)
        prio_str = f"[{prio}]" if prio < 999999 else "[?]"
        console.print(f"  {prio_str} {base}")
    console.print()

    # ── Dosyaları sırayla işle ───────────────────────────────────────────────
    basarili = 0
    basarisiz = 0
    toplam_rezerve = 0
    toplam_eksik = 0

    for i, src_path in enumerate(all_xlsx, 1):
        fname = os.path.basename(src_path)
        console.print(f"[cyan]⏳ [{i}/{len(all_xlsx)}] İşleniyor:[/cyan] {fname}")

        try:
            out_path, stats = reserve_single_file(src_path, stok_dict, output_suffix="_")

            if not out_path:
                uyarilar = stats.get("uyari", [])
                for u in uyarilar:
                    console.print(f"  [yellow]⚠️  {u}[/yellow]")
                basarisiz += 1
                continue

            toplam_rezerve += stats.get("rezerve_yapilan", 0)
            toplam_eksik += stats.get("eksik_olan", 0)

            rezerve_str = f"[green]{stats.get('rezerve_yapilan', 0)} tam/kısmi rezerve[/green]"
            eksik_str = (
                f", [red]{stats.get('eksik_olan', 0)} eksik[/red]"
                if stats.get("eksik_olan", 0) > 0
                else ""
            )
            console.print(f"  ✅ {rezerve_str}{eksik_str} → [bold]{os.path.basename(out_path)}[/bold]")

            for u in stats.get("uyari", []):
                console.print(f"  [yellow]⚠️  {u}[/yellow]")

            basarili += 1

        except Exception as exc:
            console.print(f"  [red]❌ İşlem hatası:[/red] {exc}")
            basarisiz += 1

    # ── Özet rapor ───────────────────────────────────────────────────────────
    console.print()
    from rich.table import Table as RTable
    ozet = RTable(show_header=False, box=None)
    ozet.add_column("K", style="cyan")
    ozet.add_column("V", style="magenta")
    ozet.add_row("İşlenen dosya", str(basarili))
    if basarisiz:
        ozet.add_row("Atlanan dosya", str(basarisiz))
    ozet.add_row("Toplam rezerve edilen satır", str(toplam_rezerve))
    ozet.add_row("Toplam eksik kalan satır", str(toplam_eksik))

    panel = Panel(
        ozet,
        title="[bold green]Stok Rezervasyon Tamamlandı[/bold green]",
        border_style="green",
        expand=False,
    )
    console.print(panel)
    console.print(f"\n[bold yellow]📁 Çıktılar şu klasöre kaydedildi:[/bold yellow] {path}\n")


@app.command()

def agac(
    path: str = typer.Argument(..., help="Makine Reçetesi Excel dosyası veya klasörü"),
) -> None:
    """Makine reçetesini okuyup interaktif bir Web Ağacı (HTML) oluşturur.
    Tek dosya veya klasör verilebilir. Klasör verilince içindeki tüm .xlsx dosyaları
    okunur ve tek bir HTML içinde select box ile seçilebilir hale getirilir."""
    from recipe_automation.services.bom_tree import build_bom_tree

    db_dir = os.path.join(os.getcwd(), settings.db_dir_name)
    prio_mapping = load_priority_mapping(db_dir)

    def get_file_priority_and_name(fpath, mapping):
        base = os.path.splitext(os.path.basename(fpath))[0]
        clean_name = base
        for prefix in ["Filtered_TIM_", "Filtered_", "TIM_"]:
            if clean_name.startswith(prefix):
                clean_name = clean_name[len(prefix) :]
                break
        prio = 999999
        if clean_name in mapping:
            prio = mapping[clean_name]
        else:
            for k, v in mapping.items():
                if k in clean_name or clean_name in k:
                    if v < prio:
                        prio = v
        return prio, clean_name, base

    def sort_key(fpath):
        prio, _, base = get_file_priority_and_name(fpath, prio_mapping)
        return (prio, base.lower())

    # Dosya mı klasör mü?
    if os.path.isfile(path):
        excel_files = [path]
        output_dir = os.path.dirname(path)
        output_name = f"Makine_Agaci_{os.path.splitext(os.path.basename(path))[0]}.html"
    elif os.path.isdir(path):
        raw_files = [
            os.path.join(path, f)
            for f in os.listdir(path)
            if f.lower().endswith(".xlsx") and not f.startswith("~")
        ]

        # Filtreleme çıktısı olan dosyaları (Filtered_ ile başlayanlar veya grup adı ile başlayanlar) atla
        group_names = []
        operasyon_file = os.path.join(os.getcwd(), settings.db_dir_name, "operasyon_gruplari.json")
        if os.path.exists(operasyon_file):
            try:
                with open(operasyon_file, encoding="utf-8") as f:
                    groups_data = json.load(f)
                    group_names = [str(k).upper().strip() for k in groups_data.keys()]
            except Exception:
                pass
        if not group_names:
            group_names = ["TIM", "BUKUM", "KAYNAK", "LAZER KESIM", "3D YAZICI"]

        filtered_raw_files = []
        for fpath in raw_files:
            fname = os.path.basename(fpath)
            if fname.lower().startswith("filtered_"):
                continue
            is_output = False
            for g in group_names:
                if fname.upper().startswith(g + "_"):
                    is_output = True
                    break
            if is_output:
                continue
            filtered_raw_files.append(fpath)

        excel_files = sorted(filtered_raw_files, key=sort_key)
        output_dir = path
        output_name = "Makine_Agaci_Receteler.html"
        if not excel_files:
            console.print(f"[red]Hata:[/red] Klasörde hiç .xlsx dosyası bulunamadı: {path}")
            raise typer.Exit(1)
    else:
        console.print(f"[red]Hata:[/red] {path} geçerli bir dosya veya klasör değil.")
        raise typer.Exit(1)

    console.print("\n[bold cyan]=== REÇETE AĞACI OLUŞTURULUYOR ===[/bold cyan]")
    console.print(f"[cyan]{len(excel_files)} dosya işlenecek.[/cyan]\n")

    # Her dosyayı oku ve ağaç oluştur
    all_trees = []  # [{"name": "dosya_adi", "tree": {...}}]
    for fpath in excel_files:
        fname = os.path.basename(fpath)
        console.print(f"[cyan]Dosya okunuyor:[/cyan] {fname}")

        # Doğru header satırını otomatik tespit et
        df = None
        for h in [0, 1, 2, 3]:
            try:
                candidate = pd.read_excel(fpath, header=h, nrows=5)
                cols_norm = [
                    str(c)
                    .strip()
                    .lower()
                    .replace("ı", "i")
                    .replace("ş", "s")
                    .replace("ç", "c")
                    .replace("ğ", "g")
                    for c in candidate.columns
                ]
                if any(c in ["sira", "sira no"] for c in cols_norm):
                    df = pd.read_excel(fpath, header=h)
                    break
            except Exception:
                continue

        if df is None or df.empty:
            console.print(
                f"  [yellow]Uyari:[/yellow] {fname} okunamadi veya recete sutunu bulunamadi, atlaniyor."
            )
            continue
        try:
            prio, _, base_name = get_file_priority_and_name(fpath, prio_mapping)
            displayName = base_name
            if prio < 999999:
                displayName = f"[{prio}] {base_name}"

            tree_dict = build_bom_tree(df, filename=displayName)

            # Makine adını çekip displayName'e ekle
            machine_name = ""
            if tree_dict:
                ad_val = str(tree_dict.get("ad", "")).strip()
                if ad_val and ad_val.upper() not in [
                    "ROOT",
                    "ANA MAKİNE",
                    displayName.upper(),
                    "ANA MAKINE",
                ]:
                    machine_name = ad_val
                else:
                    kod_val = str(tree_dict.get("kod", "")).strip()
                    if kod_val and kod_val.upper() not in ["ROOT"]:
                        machine_name = kod_val
            if machine_name:
                displayName = f"{displayName} ({machine_name})"

            all_trees.append({"name": displayName, "tree": tree_dict})
            console.print("  [green]OK[/green] Agac olusturuldu.")
        except Exception as e:
            console.print(f"  [yellow]Uyari:[/yellow] {fname} agac olusturulamadi: {e}")
            continue

    if not all_trees:
        console.print("[red]Hata:[/red] Hiçbir dosyadan ağaç oluşturulamadı.")
        raise typer.Exit(1)

    # JSON yapısına çevir
    all_trees_json = json.dumps(all_trees, ensure_ascii=False)

    # HTML şablonunu oku
    template_path = os.path.join(os.path.dirname(__file__), "web", "template.html")
    if not os.path.exists(template_path):
        console.print(f"[red]Hata:[/red] HTML şablonu bulunamadı: {template_path}")
        raise typer.Exit(1)

    with open(template_path, encoding="utf-8") as f:
        html_content = f.read()

    # Placeholders değiştir
    # Çoklu mod: __TREE_DATA_PLACEHOLDER__ yerine ALL_TREES_PLACEHOLDER kullan
    html_content = html_content.replace(
        "__TREE_DATA_PLACEHOLDER__", json.dumps(all_trees[0]["tree"], ensure_ascii=False)
    )
    html_content = html_content.replace("__ALL_TREES_PLACEHOLDER__", all_trees_json)
    html_content = html_content.replace("__FILE_NAME_PLACEHOLDER__", all_trees[0]["name"])

    # Çıktı dosyasını kaydet
    out_path = os.path.join(output_dir, output_name)

    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html_content)

    console.print(f"\n[green]>> Basarili![/green] {len(all_trees)} recete agaci olusturuldu.")
    console.print(
        f"[bold yellow]>>> Lutfen su dosyaya cift tiklayarak acin:[/bold yellow] {out_path}\n"
    )

    # Otomatik tarayıcıda aç
    try:
        import webbrowser

        webbrowser.open(out_path)
    except:
        pass


if __name__ == "__main__":
    app()
